
import streamlit as st
import pandas as pd
import io
import warnings
import numpy as np
from numpy.linalg import lstsq
from datetime import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
from tqdm import tqdm
import plotly.express as px
import seaborn as sns
from scipy.optimize import minimize
import openpyxl
from openpyxl.drawing.image import Image
from xlsxwriter import Workbook
import matplotlib.ticker as ticker
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from urllib.parse import urljoin, quote
# import matplotlib.font_manager as fm
# import tensorflow as tf
# from keras import backend as K
# from keras.models import Sequential
# from keras.layers import Dense, LSTM
# from sklearn.preprocessing import MinMaxScaler
import pickle
# import gc
from functools import lru_cache

warnings.filterwarnings('ignore')

plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False  # 마이너스 부호 깨짐 방지

for k in ["plots", "params", "quantile", "real", "combined", "combined_all","combined_i", "moving_avg", "simulated_paths", "lower_bound",
          "simulated_paths_i", "upper_bound", "mean_rates", "std_dev_rates"]:
    st.session_state.setdefault(k, {})

# y 축 틱 레이블을 백만 원 단위로 설정
def millions_formatter(x, pos):
    return f'{x / 1_000_000:,.0f}'

# ---------- Vasicek Calibration (AR(1) 등가 OLS) ----------
def vasicek_calibrate(rates: pd.Series, dt=1/12):
    r = np.asarray(rates, dtype=float)
    x = r[:-1]
    y = r[1:]
    # y = a + b*x + e  → b = exp(-kappa*dt), a = theta*(1 - b)
    X = np.column_stack([np.ones_like(x), x])  # [1, x]
    beta, _, _, _ = lstsq(X, y, rcond=None)
    a_hat, b_hat = beta
    # 방어: b_hat이 (0,1) 밖이면 클리핑
    b_hat = np.clip(b_hat, 1e-6, 1 - 1e-6)
    kappa = -np.log(b_hat) / dt
    theta = a_hat / (1 - b_hat)
    # 잔차로 sigma 추정
    resid = y - (a_hat + b_hat * x)
    sigma_eps = np.std(resid, ddof=1)
    # eps ~ N(0, sigma_eps^2) 이고, 연속시간 sigma는:
    sigma = sigma_eps * np.sqrt(2 * kappa / (1 - np.exp(-2 * kappa * dt)))
    r0 = r[-1]
    return kappa, theta, sigma, r0


# ---------- Vasicek Simulation (Euler) ----------
def vasicek_simul(r0, kappa, theta, sigma, simulation_times, T, dt, epsilon_var):
    # --- scalarize (핵심) ---
    r0    = float(np.asarray(r0).squeeze())
    kappa = float(np.asarray(kappa).squeeze())
    theta = float(np.asarray(theta).squeeze())   # ★ 여기 때문에 에러 났던 것
    sigma = float(np.asarray(sigma).squeeze())
    dt    = float(np.asarray(dt).squeeze())

    eps = np.asarray(epsilon_var, dtype=np.float64)
    if eps.shape != (simulation_times, T):
        raise ValueError(f"epsilon_var shape mismatch: expected {(simulation_times, T)}, got {eps.shape}")

    out = np.zeros((simulation_times, T), dtype=np.float64)
    out[:, 0] = r0
    sdt = np.sqrt(dt)

    for i in range(simulation_times):
        for j in range(1, T):
            dW = float(eps[i, j])  # 스칼라 보장
            out[i, j] = out[i, j-1] + kappa*(theta - out[i, j-1])*dt + sigma*sdt*dW

    return out


# ---------- Vasicek Objective (평균레벨 맞추기) ----------
def vasicek_objective(kappa, theta, sigma, r0, simulation_times, T, dt, epsilon_var, target_level):
    kappa = float(np.maximum(kappa[0], 1e-6))
    sims = vasicek_simul(r0, kappa, theta, sigma, simulation_times, T, dt, epsilon_var)
    # 방법 A: 마지막 시점 평균레벨을 타깃으로
    mean_terminal = sims[:, -1].mean()
    return (mean_terminal - target_level)**2


# ---------- GBM Simulation ----------
def gbm_simul(S0, mu, sigma, simulation_times, T, dt, epsilon_var):
    """
    dS/S = mu*dt + sigma*dW  → S_t = S0 * exp( (mu - 0.5*sigma^2)*t + sigma*W_t )
    epsilon_var: (simulation_times x T) dW ~ N(0,1)
    """
    eps = np.asarray(epsilon_var, dtype=float)  # shape: (simulation_times, T)
    assert eps.shape == (simulation_times, T), "epsilon_var shape mismatch"

    out = np.zeros((simulation_times, T), dtype=float)
    out[:, 0] = S0
    sdt = np.sqrt(dt)
    drift = (mu - 0.5 * sigma**2) * dt

    for i in range(simulation_times):
        logS = np.log(S0)
        for j in range(1, T):
            dW = eps[i, j] * sdt
            logS = logS + drift + sigma * dW
            out[i, j] = np.exp(logS)
    return out


# ---------- GBM Objective (목표 수익률 매칭) ----------
def gbm_objective(mu, target_annual_return, S0, sigma, simulation_times, T, dt, epsilon_var):
    mu = float(mu[0])  # 월 로그수익률
    sims = gbm_simul(S0, mu, sigma, simulation_times, T, dt, epsilon_var)
    mean_terminal = sims[:, -1].mean()

    # 목표지수 수준: 연 수익률 target → 월 로그수익률로 변환
    monthly_target_return = (1 + target_annual_return) ** (1/12) - 1
    target_log_ret = np.log(1 + monthly_target_return)
    target_level = S0 * np.exp(target_log_ret * T)

    return (mean_terminal - target_level)**2


# ---------- Wrapper: find_optimal_mu ----------
def find_optimal_mu(target_annual_return, S0, sigma, simulation_times, T, dt, epsilon_var, init_mu=None):
    if init_mu is None:
        init_mu = np.log(1 + target_annual_return) / 12  # 초기 월 로그수익률 추정
    res = minimize(
        gbm_objective, x0=init_mu,
        args=(target_annual_return, S0, sigma, simulation_times, T, dt, epsilon_var),
        method="Nelder-Mead"
    )
    return float(res.x[0])


def VasicekCalibration(rates, dt=1/12):
    n = len(rates)
    # Implement MLE to calibrate parameters     
    Sx = sum(rates[0:(n-1)])
    Sy = sum(rates[1:n])
    Sxx = np.dot(rates[0:(n-1)], rates[0:(n-1)])
    Sxy = np.dot(rates[0:(n-1)], rates[1:n])
    Syy = np.dot(rates[1:n], rates[1:n])
    theta = (Sy * Sxx - Sx * Sxy) / (n * (Sxx - Sxy) - (Sx**2 - Sx*Sy))
    kappa = -np.log((Sxy - theta * Sx - theta * Sy + n * theta**2) / (Sxx - 2*theta*Sx + n*theta**2)) / dt
    a = np.exp(-kappa * dt)
    sigmah2 = (Syy - 2*a*Sxy + a**2 * Sxx - 2*theta*(1-a)*(Sy - a*Sx) + n*theta**2 * (1-a)**2) / n
    sigma = np.sqrt(sigmah2*2*kappa / (1-a**2))
    r0 = rates[n-1]
    return [kappa, theta, sigma, r0]

def VasicekSimul(var, kappa, mean, sd, simulation_times, dt):    
    rates = macro[var]
    theta = mean
    sigma = sd
    r0 = VasicekCalibration(rates, dt=1/12)[3]
    results = np.zeros((simulation_times, T))
    results[:, 0] = r0  # 초기 조건 설정

    for i in range(simulation_times):
        results[i, 0] = rates.iloc[-1] + kappa * (theta - rates.iloc[-1]) * dt + sigma * all_epsilon[var].iloc[i,0]
        for j in range(1, T):
            dW = all_epsilon[var].iloc[i,j]
            results[i, j] = results[i, j-1] + kappa * (theta - results[i, j-1]) * dt + sigma * dW
    return results

def VasicekSimul_Shifted(var, kappa, mean, sd, simulation_times, dt, shift_constant):    
    rates = macro[var] + shift_constant
    theta = mean + shift_constant
    sigma = sd
    results = np.zeros((simulation_times, T))
    results[:, 0] = rates.iloc[-1]  # 초기 조건 설정

    for i in range(simulation_times):
        drift = kappa * (theta - results[i, 0]) * dt
        diffusion = sigma * all_epsilon[var].iloc[i,0]
        results[i, 0] += drift + diffusion
        
        for j in range(1, T):
            dW = all_epsilon[var].iloc[i,j]
            drift = kappa * (theta - results[i, j-1]) * dt
            diffusion = sigma * dW
            results[i, j] = results[i, j-1] + drift + diffusion
    
    return results - shift_constant # 최종 결과에서 shift를 다시 제거    

def kappa_objective(kappa, *args):
    var, simulation_times, dt, mean, sd = args
    simulated_paths = VasicekSimul(var, kappa, mean, sd, simulation_times, dt)
    return (np.mean(simulated_paths) - mean)**2

def kappa_objective_shifted(kappa, *args):
    var, simulation_times, dt, mean, sd, shift_constant = args  # Add shift_constant to arguments
    simulated_paths = VasicekSimul_Shifted(var, kappa, mean, sd, simulation_times, dt, shift_constant)
    return (np.mean(simulated_paths) - (mean + shift_constant))**2

def GBMSimul(var, mean, sd, simulation_times, dt):    
    rates = macro[var]
    results = np.zeros((simulation_times, T))

    for i in range(simulation_times):
        results[i, 0] = rates.iloc[-1] + (mean - (sd ** 2) / 2) * dt + sd * all_epsilon[var].iloc[i,0] * (dt ** 0.5)
        for j in range(1, T):
            dW = all_epsilon[var].iloc[i,j]
            results[i, j] = results[i, j-1] + (mean - (sd ** 2) / 2) * dt + sd * dW * (dt ** 0.5)
    return results

def GBM_objective(mean, target_mean, var, sd, simulation_times, dt):
    simulated_paths = GBMSimul(var, mean, sd, simulation_times, dt)
    return (np.mean(simulated_paths) - target_mean)**2

def find_optimal_mean(target_mean, var, initial_mean, sd, simulation_times, dt):
    args = (var, sd, simulation_times, dt)  # initial_mean 제거
    result = minimize(GBM_objective, initial_mean, args=(target_mean, *args))
    return result.x[0]

def create_dataset(dataset, time_step=1, target_idx = 0):  # target_column은 G_Bond_10Y를 나타냅니다.
    dataX, dataY = [], []
    for i in range(len(dataset) - time_step - 1):
        a = dataset[i:(i + time_step), :]
        dataX.append(a)
        dataY.append(dataset[i + time_step, target_idx])  # G_Bond_10Y 특성을 선택합니다.
    return np.array(dataX), np.array(dataY)

def predict_future(model, input_data, steps, scaler):
    var_num = input_data.shape[1]
    input_data = np.reshape(input_data, (1, input_data.shape[0], input_data.shape[1]))
    future_predictions = []  # 미래 예측을 저장할 리스트
    
    # 한 단계씩 예측을 수행
    for _ in range(steps):
        prediction = model.predict(input_data)  # 현재 입력 데이터로 예측 수행
        future_predictions.append(prediction[0, 0])  # 예측 값을 리스트에 추가
        
        # 예측된 값을 입력 데이터의 끝에 추가하고, 가장 오래된 데이터를 제거
        new_input_data = np.append(input_data[0, 1:], [[prediction[0, 0]] + list(input_data[0, -1, 1:var_num])], axis=0)
        input_data = np.reshape(new_input_data, (1, new_input_data.shape[0], new_input_data.shape[1]))
    
    future_predictions_dummy = np.hstack((np.array(future_predictions).reshape(-1, 1), np.zeros((len(future_predictions), var_num - 1))))
    future_predictions = scaler.inverse_transform(future_predictions_dummy)[:, 0]
    return future_predictions

# def LSTM_simul(asset, seed):
#     data = macro[asset_related[asset]].values
    
#     # 데이터 스케일링
#     scaler = MinMaxScaler(feature_range=(0, 1))
#     data_scaled = scaler.fit_transform(data)

#     # 학습 데이터와 테스트 데이터 준비
#     target_idx = 0
#     time_step = 12
#     train_size = int(len(data_scaled) * 0.75)
#     train, test = data_scaled[0:train_size, :], data_scaled[train_size:len(data_scaled), :]
#     X_train, y_train = create_dataset(train, time_step, target_idx=target_idx)
#     X_test, y_test = create_dataset(test, time_step, target_idx=target_idx)

#     tf.random.set_seed(seed)
#     # LSTM 모델 생성
#     model = Sequential()
#     model.add(LSTM(50, return_sequences=True, input_shape=(X_train.shape[1], X_train.shape[2])))
#     model.add(LSTM(50, return_sequences=True))
#     model.add(LSTM(50))
#     model.add(Dense(1))
#     model.compile(loss='mean_squared_error', optimizer='adam')

#     # 모델 학습
#     model.fit(X_train, y_train, validation_data=(X_test, y_test), epochs=100, batch_size=64, verbose=0)

#     #예측
#     train_predict_dummy = np.hstack((model.predict(X_train), np.zeros((X_train.shape[0], data.shape[1] - 1))))
#     test_predict_dummy = np.hstack((model.predict(X_test), np.zeros((X_test.shape[0], data.shape[1] - 1))))

#     train_predict = scaler.inverse_transform(train_predict_dummy)[:, target_idx]
#     test_predict = scaler.inverse_transform(test_predict_dummy)[:, target_idx]

#     train_x_values = np.arange(time_step, len(train_predict) + time_step)
#     test_x_values = np.arange(len(train_predict) + 2*time_step, len(train_predict) + len(test_predict) + 2*time_step)

#     last_test_data = X_test[-1]
#     future_predictions = predict_future(model, last_test_data, steps = 12 * projection_period, scaler=scaler)
    
#     return data, train_predict, test_predict, train_x_values, test_x_values, future_predictions, time_step

# info_dict
@lru_cache(maxsize=None)
def get_제도설계_info(가입대상분류):
    return st.session_state.제도설계.loc[st.session_state.제도설계['직급'] == 가입대상분류].iloc[0]

@lru_cache(maxsize=None)
def get_기초율_multiplier(age1, age2, info):
    multiplier = 1.0
    for idx, row in 기초율.iterrows():
        age = row['연령']
        if age1 <= age < age2:
            multiplier *= (1 + row[info])
    return multiplier

def 나이계산(birthdate, reference_date):
    delta = relativedelta(reference_date, birthdate)
    return int(round(delta.years + delta.months/12 + delta.days/365.25, 0))

def get_근속년수_round(지급률_계산, 시산기준일, 기산일):
    delta = relativedelta(시산기준일, 기산일)
    delta += relativedelta(days=1)
    years = delta.years
    months = delta.months
    days = delta.days

    if 지급률_계산 == '일할':
        # Calculate the difference in days
        근속년수 = years + months / 12 + days / 365.25

    elif 지급률_계산 == '월 절상':
        if days > 0:
            days = 0
            months = months + 1
            근속년수 = years + months / 12
        else:
            days = 0
            months = months
            근속년수 = years + months / 12

    elif 지급률_계산 == '반기 절상':
        if days > 0 and months < 6:
            days = 0
            months = 6
            years = years
            근속년수 = years + 0.5
        elif days > 0 and months >= 6:
            days = 0
            months = 0
            years = years + 1
            근속년수 = years
        else:
            days = 0
            months = 0
            근속년수 = years

    return 근속년수

def info_dict(명부, mp, 시산기준일, bu_i):
    info = {}
    info['사번'] = 명부['사원번호'].iloc[mp]
    info['성별'] = 1 if 명부['성별'].iloc[mp] == 'M' else 2
    info['현재연령'] = 명부['연령'].iloc[mp]
    info['시산연령'] = 명부['연령'].iloc[mp] + 시산기준일.year - max(명부['입사일'].iloc[mp].year, 기준일.year)
    info['가산개수'] = 명부['가산개수'].iloc[mp]
    info['근속년수'] = get_근속년수_round(get_제도설계_info(명부['가입대상분류'].iloc[mp])['지급률_계산'], 시산기준일, 명부['기산일'].iloc[mp])
    info['평가방법'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['평가방법']
    info['정년초과_평가'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년초과_평가']
    info['정년초과_정년'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년초과_정년']
    info['지급률_계산'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['지급률_계산']

    if get_제도설계_info(명부['가입대상분류'].iloc[mp])['지급률_설정'] == '지급률 테이블':
        info['지급률'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['지급률']
        info['bx'] = 지급률.iloc[지급률.index[지급률['근속년수'] == int(info['근속년수'])], 지급률.columns.get_loc(info['지급률']) - 1].values[0]
        info['bx_누적'] = 지급률.iloc[지급률.index[지급률['근속년수'] == int(info['근속년수'])], 지급률.columns.get_loc(info['지급률'])].values[0] + (info['근속년수'] - int(info['근속년수'])) * info['bx']        

    elif get_제도설계_info(명부['가입대상분류'].iloc[mp])['지급률_설정'] == '명부 입력':
        info['지급률'] = '명부 입력'
        info['bx'] = 명부['지급배수'].iloc[mp]
        info['bx_누적'] = info['근속년수'] * 명부['지급배수'].iloc[mp]

    #info['bx'] = 지급률.iloc[지급률.index[지급률['근속년수'] == int(info['근속년수'])], 지급률.columns.get_loc(info['지급률']) - 1].values[0]
    #info['bx_누적'] = 지급률.iloc[지급률.index[지급률['근속년수'] == int(info['근속년수'])], 지급률.columns.get_loc(info['지급률'])].values[0] + (info['근속년수'] - int(info['근속년수'])) * info['bx']
    info['할당'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['할당']
    info['임직원'] = 명부['가입대상분류'].iloc[mp]
    info['bu'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['bu'] + bu_i
    info['승급률'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['승급률']
    info['퇴직률'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['퇴직률']
    info['사망률'] = '표준사망률'
    #info['할인율'] = 할인율
    info['추계액_증분'] = 0 if info['평가방법'] == 'PUC' else 명부['임원연말퇴직금추계액'].iloc[mp] - 명부['당월퇴직추계'].iloc[mp]

    #정년: 정년초과자는 다음해 퇴직 OR 추계액 방식 설정
    if  get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년_설정'] == '현재연령 기준':
        info['정년'] = info['현재연령'] + get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년']
        #info['정년'] = info['현재연령'] + int(re.search(r'\d+', get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년']).group())
    else:
        info['정년'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년']

    if info['현재연령'] > info['정년']:
        info['평가방법'] = info['정년초과_평가']
        
        if get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년초과_설정'] == '현재연령 기준':
            info['정년'] = info['현재연령'] + get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년초과_정년'] - 1
        else:
            info['정년'] = get_제도설계_info(명부['가입대상분류'].iloc[mp])['정년초과_정년']

    #기준급여: 시산년도 != 현재연도 -> 갭 year 간의 (승급률 + bu)반영
    multiplier = get_기초율_multiplier(info['현재연령'], info['시산연령'], info['승급률'])
    info['기준급여'] = 명부['기준급여'].iloc[mp] * np.power(1 + info['bu'], 시산기준일.year - 기준일.year) * multiplier

    # 추계액(미래 추계액은 법정제로 계산됨(기준급여 * 근속년수))
    if 기준일 == 시산기준일:
        info['추계액'] = 명부['당월퇴직추계'].iloc[mp]
    else: 
        info['추계액'] = info['추계액'] = info['기준급여'] * (info['bx_누적'] + info['가산개수'])

    return info

def lapse(info_dict):
    lapse_table = pd.DataFrame(columns = ['현재연도','현재연령','잔존율_보정','탈퇴계수(누적)','탈퇴계수'])

    lapse_table['현재연령'] = range(info_dict['현재연령'], info_dict['정년'] + 1)
    lapse_table['현재연도'] = range(기준일.year, 기준일.year + info_dict['정년'] - info_dict['현재연령'] + 1)

    lapse_table.at[0, '잔존율_보정'] = 1
    lapse_table.at[0, '탈퇴계수(누적)'] = 0
    lapse_table.at[0, '탈퇴계수'] = 0

    if info_dict['평가방법'] == 'PUC':
        for i in range(1, len(lapse_table)):
            age = lapse_table.at[i - 1, '현재연령']
            퇴직률 = 기초율.loc[기초율['연령'] == age, info_dict['퇴직률']].values[0]
            사망률 = 기초율.iloc[기초율.index[기초율['연령'] == age], info_dict['성별'] + 8].values[0]
            lapse_table.at[i, '잔존율_보정'] = lapse_table.at[i - 1, '잔존율_보정'] * (1 - 퇴직률 - 사망률)
            lapse_table.at[i, '탈퇴계수(누적)'] = 1 - lapse_table.at[i, '잔존율_보정']
            lapse_table.at[i, '탈퇴계수'] = lapse_table.at[i, '탈퇴계수(누적)'] - lapse_table.at[i - 1, '탈퇴계수(누적)']
    #추계액 방식은 탈퇴계수 0으로 미래 부채에 보정 없음
    else:
        lapse_table['탈퇴계수(누적)'] = 0
        lapse_table['탈퇴계수'] = 0
    
    return lapse_table

def DBO(info_dict, 할인율):
    
    if info_dict['평가방법'] == 'PUC':
        if info_dict['시산연령'] <= info_dict['정년']:
            cal_table = np.zeros((info_dict['정년'] - info_dict['시산연령'] + 1, 19))
            cal_table[:, 1] = np.arange(info_dict['시산연령'], info_dict['정년'] + 1)
        else:
            cal_table = np.zeros((1,19))
            cal_table[:, 1] = info_dict['시산연령']
            
    else: 
        cal_table = np.zeros((1,19))
        cal_table[:, 1] = info_dict['시산연령']
    
    cal_table[:, 0] = 시산기준일.year + np.arange(len(cal_table))
    cal_table[:, 2] = np.arange(len(cal_table))
    cal_table[:, 5] = info_dict['근속년수'] + np.arange(len(cal_table))
    
    #6 지급률(Bx)
    if info_dict['지급률'] == '명부 입력':
        cal_table[:, 6] = cal_table[:, 5] * info_dict['bx'] + info_dict['가산개수'] #지급률에 가산개수 더해줌 -> 가산개수 있으면 할당을 지급률 할당으로!
    else:
        bx_values = np.zeros(len(cal_table))
        bx_indices = np.searchsorted(지급률['근속년수'], cal_table[:, 5].astype(int))
        for i, bx_idx in enumerate(bx_indices):
            bx_values[i] = 지급률.loc[bx_idx, info_dict['지급률']]
            bx_values[i] += (cal_table[i, 5] - int(cal_table[i, 5])) * 지급률.iloc[bx_idx, 지급률.columns.get_loc(info_dict['지급률']) - 1]
        cal_table[:, 6] = bx_values + info_dict['가산개수'] #지급률에 가산개수 더해줌 -> 가산개수 있으면 할당을 지급률 할당으로!
    
    #3 임금상승률(bx)
    age_values = 기초율.loc[기초율['연령'].isin(cal_table[:, 1]), info_dict['승급률']].values
    age_indices = np.searchsorted(cal_table[:, 1], 기초율['연령'].values)
    cal_table[:, 3] = (1 + age_values) * (1 + info_dict['bu'])

    #4 기준급여(Sx)
    cal_table[:, 4] = info_dict['기준급여']
    for i in range(1, len(cal_table)):
        cal_table[i, 4] = cal_table[i - 1, 4] * cal_table[i - 1, 3]
    

    #8 퇴직률(qx_w)
    qx_w_values = 기초율.loc[기초율['연령'].isin(cal_table[:, 1]), info_dict['퇴직률']].values
    qx_w_indices = np.searchsorted(cal_table[:, 1], 기초율['연령'].values)
    cal_table[:, 8] = qx_w_values
    cal_table[-1, 8] = 1

    #9 사망률(qx_d)
    qx_d_values = np.zeros(len(cal_table))
    qx_d_indices = np.searchsorted(기초율['연령'], cal_table[:, 1])

    for i, qx_d_idx in enumerate(qx_d_indices):
        qx_d_values[i] = 기초율.iloc[qx_d_idx, info_dict['성별'] + 8]
    cal_table[:, 9] = qx_d_values
    cal_table[-1, 9] = 0


    #7 생존률(px)
    cal_table[:, 7] = 1
    for i in range(1, len(cal_table)):
        cal_table[i, 7] = cal_table[i - 1, 7] * (1 - cal_table[i - 1, 8] - cal_table[i - 1, 9])

    #10 퇴직급여_기간안분(기시)
    if info_dict['할당'] == '근속년수':
        cal_table[:, 10] = cal_table[:, 4] * cal_table[:, 6] * cal_table[0, 6] / cal_table[:, 6]
    elif info_dict['할당'] == '지급률':     
        cal_table[:, 10] = cal_table[:, 4] * cal_table[:, 6] * cal_table[0, 5] / cal_table[:, 5]

    #11 퇴직급여
    cal_table[:-1, 11] = (cal_table[:-1, 10] + cal_table[1:, 10]) * 0.5
    cal_table[-1, 11] = cal_table[-1, 10]

    #12 할인율
    cal_table[:-1, 12] = np.power(1 / (1 + 할인율), cal_table[:-1, 2] + 0.5)
    cal_table[-1, 12] = np.power(1 / (1 + 할인율), cal_table[-1, 2])

    #13 DBO
    if info_dict['평가방법'] == 'PUC':  #[11]: 퇴직급여_기간안분(기말 + 기시 / 2) | [7]:px  | [8]: qx [9]: dx
        cal_table[:, 13] = cal_table[:, 11] * cal_table[:, 7] * (cal_table[:, 8] + cal_table[:, 9]) * cal_table[:, 12]
    else:
        cal_table[:, 13] = info_dict['추계액']
    if cal_table[0, 5] < 1: #근속년수 1년 미만은 DBO_t0 = 0
        cal_table[0, 13] = 0

    #14 t * DBO
    if info_dict['평가방법'] == 'PUC':
        cal_table[:-1, 14] = cal_table[:-1, 13] * (cal_table[:-1, 2] + 0.5)
        cal_table[-1, 14] = cal_table[-1, 13] * cal_table[-1, 2]
    else:
        cal_table[:, 14] = cal_table[:, 13] * (info_dict['정년'] - cal_table[:, 1])

    #15 퇴직급여_기간안분(기말)   [4]: 지급률 | [6]: 기준급여
    cal_table[0, 15] = cal_table[0, 10]
    if cal_table.shape[0] > 1:
        cal_table[1:, 15] = cal_table[1:, 4] * cal_table[1:, 6] * cal_table[1, 6] / cal_table[1:, 6]

    #16 NC
    if info_dict['평가방법'] == 'PUC':  # [15]: 퇴직급여_기간안분(기말) | [10]: 퇴직급여_기간안분(기시) | [7]:px  | [8]: qx | [9]: dx | [12]: 할인율
        nc_values = ((cal_table[:-1, 15] + cal_table[1:, 15]) * 0.5 - (cal_table[:-1, 10] + cal_table[1:, 10]) * 0.5) * cal_table[:-1, 7] * (cal_table[:-1, 8] + cal_table[:-1, 9]) * cal_table[:-1, 12]
        cal_table[:-1, 16] = nc_values
        cal_table[-1, 16] = (cal_table[-1, 15] - cal_table[-1, 10]) * cal_table[-1, 7] * (cal_table[-1, 8] + cal_table[-1, 9]) * cal_table[-1, 12]
    else:
        if 기준일 == 시산기준일:
            cal_table[0, 16] = info_dict['추계액_증분']
        else:
            cal_table[0, 16] =  info_dict['bx'] * info_dict['기준급여']
    if cal_table[0, 5] < 1: #임원인데 차년도 추계액 없거나, 근속년수 1년 미만이면 NC_t0 = 0
        cal_table[0, 16] = 0

    #17 추계액
    if 기준일 == 시산기준일:
        cal_table[0, 17] = info_dict['추계액']
        if cal_table.shape[0] > 1:
            cal_table[1:, 17] = (cal_table[1:, 6]) * cal_table[1:, 4]   #[4]: 기준급여 [6]: 지급률
    else:
        cal_table[0, 17] = (cal_table[0, 6]) * cal_table[0, 4]
        if cal_table.shape[0] > 1:
            cal_table[1:, 17] = (cal_table[1:, 6]) * cal_table[1:, 4]

    #18 EBP #[17]: 기준급여 * 지급률 | [7]:px  | [8]: qx [9]: dx
    cal_table[:, 18] = 0
    if info_dict['시산연령'] == info_dict['정년']:
        cal_table[0, 18] = cal_table[0, 17]

    if info_dict['평가방법'] == 'PUC':
        cal_table[:-1, 18] = (cal_table[:-1, 17] + cal_table[1:, 17]) * 0.5 * cal_table[:-1, 7] * (cal_table[:-1, 8] + cal_table[:-1, 9]) 
        cal_table[-1, 18] = cal_table[-1, 17] * cal_table[-1, 7] * (cal_table[-1, 8] + cal_table[-1, 9])

    return cal_table
def sample_res(cal_table, lapse_table):
    sample_res = {}

    보정계수_누적 = 1 - lapse_table.loc[lapse_table['현재연령'] == cal_table[0, 1], '탈퇴계수(누적)'].values[0]
    보정계수 = 1 - lapse_table.loc[lapse_table['현재연령'] == cal_table[0, 1], '탈퇴계수'].values[0]
    sample_res['시산연령'] = cal_table[0, 1]
    sample_res['근속년수'] = cal_table[0, 5]
    sample_res['지급률'] = cal_table[0, 6]
    sample_res['추계액'] = cal_table[0, 17] * 보정계수_누적
    sample_res['DBO'] = cal_table[:, 13].sum() * 보정계수_누적
    sample_res['t*DBO'] = cal_table[:, 14].sum() * 보정계수_누적
    sample_res['DUR'] = "N/A" if sample_res['DBO'] == 0 else sample_res['t*DBO'] / sample_res['DBO']
    sample_res['NC'] = cal_table[:, 16].sum() * 보정계수_누적
    sample_res['EBP'] = cal_table[0, 18] * 보정계수_누적
    sample_res['CF'] = cal_table[:,13]
    return sample_res, 보정계수 

def simul(명부, 기준일, sim_yr, 할인율, bu_i):
    global retired_cumulative
    if sim_yr == 0:
        retired_cumulative = 0

    시산기준일 = date(기준일.year + sim_yr, 기준일.month, 기준일.day)
    total_emp = len(명부)
    result_data = np.zeros((total_emp, 22), dtype = 'object')
    cf_data = {}

    for mp in tqdm(range(total_emp), desc="Processing Employees", unit="employee"):
        info = info_dict(명부, mp, 시산기준일, bu_i)

        if info['시산연령'] > info['정년']:
            #print(f"Employee {info['사번']} has reached or exceeded 정년. Skipping calculation.")
            result_data[mp, :] = [
                info['사번'],
                info['현재연령'],
                info['시산연령'],
                0,
                0,
                0,
                0,
                0,
                0,
                0,
                0,
                0,
                0,
                info['평가방법'],
                info['정년'],
                info['할당'],
                info['임직원'],
                info['bu'],
                info['승급률'],
                info['퇴직률'],
                info['사망률'],
                0
            ]
        else:
            근속년수 = info['근속년수']
            lapse_table = lapse(info)
            cal_table = DBO(info, 할인율)
            sample_res_dict, 보정계수 = sample_res(cal_table, lapse_table)

            result_data[mp, :] = [
                info['사번'],
                info['현재연령'],
                info['시산연령'],
                근속년수,
                sample_res_dict['지급률'],
                info['기준급여'],
                sample_res_dict['DBO'],
                sample_res_dict['추계액'],
                sample_res_dict['NC'],
                sample_res_dict['EBP'],
                sample_res_dict['t*DBO'],
                sample_res_dict['DUR'],
                보정계수,
                info['평가방법'],
                info['정년'],
                info['할당'],
                info['임직원'],
                info['bu'],
                info['승급률'],
                info['퇴직률'],
                info['사망률'],
                할인율
            ]
            cf_data[mp] = sample_res_dict['CF']
    
    result_table = pd.DataFrame(result_data, columns=[
        '사번', '현재연령', '시산연령', '근속년수', '지급률', '기준급여', 'DBO', '추계액', 'NC', 'EBP', 't*DBO',
        'DUR', '보정계수', '평가방법', '정년', '할당', '임직원', 'bu', '승급률', '퇴직률', '사망률', '할인율'
    ])
    
    퇴직자 = 0 if sim_yr == 0 else retired[sim_yr - 1]
    if sim_yr > 0 and 신입가정 > 0:
        new_emp = int(np.sum(1 - result_data[:, 12][result_data[:, 12] != 0])) + 퇴직자
        new_emp_data = []
        
        # Create a DataFrame for new employees with the given assumptions
        for i in range(new_emp):
    
            # Randomly assign gender based on the given ratio (70% male, 30% female)
            if len(st.session_state.신입명부) == 1:
                new = [1] * round(new_emp * st.session_state.신입명부['비중'][1])
            else:
                new = [1] * round(new_emp * st.session_state.신입명부['비중'][1]) + [2] * (new_emp - round(new_emp * st.session_state.신입명부['비중'][1]))

            입사일구분 = st.session_state.신입명부['입사일'][new[i]]   # 연초 / 연중 / 연말
            if 입사일구분 == "연초":
                tmp = 시산기준일 - relativedelta(years=1) + relativedelta(days=1)
                entry_dt = date(tmp.year, tmp.month, tmp.day)
            elif 입사일구분 == "연중":
                tmp = 시산기준일 - relativedelta(months=6) + relativedelta(days=1)
                entry_dt = date(tmp.year, tmp.month, tmp.day)
            else:   # 연말
                entry_dt = 시산기준일

            new_emp_data.append({
                '사원번호': len(명부) + i + 1,
                '가입대상분류': st.session_state.신입명부['직급'][new[i]],
                '입사일': entry_dt, # date(시산기준일.year, st.session_state.신입명부['입사일_date'][new[i]].month, st.session_state.신입명부['입사일_date'][new[i]].day),
                '기산일': entry_dt, #date(시산기준일.year, st.session_state.신입명부['입사일_date'][new[i]].month, st.session_state.신입명부['입사일_date'][new[i]].day),
                '기준급여': st.session_state.신입명부['기준급여'][new[i]],
                '당월퇴직추계': 0,
                '임원연말퇴직금추계액': 0,
                '지급배수': 1,
                '가산개수': 0,
                #'가산년수': 0,
                '성별': st.session_state.신입명부['성별'][new[i]],
                '생년월일': 시산기준일 - pd.DateOffset(years=st.session_state.신입명부['연령'][new[i]]),
                '연령': st.session_state.신입명부['연령'][new[i]],
                '근속년수': (시산기준일 - entry_dt).days / 365.25
            })
        new_emp_df = pd.DataFrame(new_emp_data)
        명부 = pd.concat([명부, new_emp_df], ignore_index=True)
        
        result_data_ = np.zeros((total_emp + new_emp, 22), dtype='object')    
        result_data_[:total_emp, :] = result_data
        result_data = result_data_

        # Calculate the values for new employees and add them to the result_table
        for i in tqdm(range(new_emp), desc="Processing Employees", unit="employee"):
            mp = total_emp + i  # Index of the new employee in the result_data
            info = info_dict(명부, mp, 시산기준일, bu_i)
            lapse_table = lapse(info)
            cal_table = DBO(info, 할인율)
            sample_res_dict, 보정계수 = sample_res(cal_table, lapse_table)

            result_data[mp, :] = [
                info['사번'],
                info['현재연령'],
                info['시산연령'],
                info['근속년수'],
                sample_res_dict['지급률'],
                info['기준급여'],
                sample_res_dict['DBO'],
                sample_res_dict['추계액'],
                sample_res_dict['NC'],
                sample_res_dict['EBP'],
                sample_res_dict['t*DBO'],
                sample_res_dict['DUR'],
                보정계수,
                info['평가방법'],
                info['정년'],
                info['할당'],
                info['임직원'],
                info['bu'],
                info['승급률'],
                info['퇴직률'],
                info['사망률'],
                할인율
            ]
            cf_data[mp] = sample_res_dict['CF']
    
    result_table = pd.DataFrame(result_data, columns=[
        '사번', '현재연령', '시산연령', '근속년수', '지급률', '기준급여', 'DBO', '추계액', 'NC', 'EBP', 't*DBO',
        'DUR', '보정계수', '평가방법', '정년', '할당', '임직원', 'bu', '승급률', '퇴직률', '사망률', '할인율'
    ])
    current_retired = (result_table['정년'] <= result_table['시산연령']).sum()
    new_retired = current_retired - (retired_cumulative if sim_yr > 0 else 0)
    retired.append(new_retired)
    retired_cumulative = current_retired

    # result_table = result_table[result_table['정년'] >= result_table['시산연령']]
    명부 = 명부[명부['사원번호'].isin(result_table['사번'])]
    
    return result_table, 명부, cf_data
# ===== 공통 헬퍼 =====
def asset_config_block(
    title: str,
    base_default: str,
    exp_key: str,          # Expected_returns column key
    sigma_key: str,        # Expected_sigma column key
    model_default_index: int,
    related_default: list[str],
    key_prefix: str,
    use_index_suffix: bool = True,   # 기초자산에 "_index" 붙일지 여부
):
    st.write(title)
    options = macro.columns[1:]  # 후보군
    base_idx = options.get_loc(base_default) if base_default in options else 0

    # ① 기초자산(단일)
    base = st.selectbox("자산군 선택", options, index=base_idx, key=f"{key_prefix}_base")

    # ② 모델(단일)
    model = st.selectbox("시뮬레이션 모델", ['Vasicek','GBM'], #'LSTM'
                        index=model_default_index, key=f"{key_prefix}_model")

    # ③ 기대수익률
    exp_opt = st.selectbox("기대수익률",
                        ['Building Block','MLE','Vasicek','GBM','직접입력'], #'LSTM'
                        index=4, key=f"{key_prefix}_exp_opt")
    if exp_opt == '직접입력':
        exp_val = st.number_input("기대수익률(직접입력)",
                                value=float(Expected_returns.loc["Expected_returns", exp_key]),
                                step=0.0001, format="%.4f", key=f"{key_prefix}_exp_val")
    else:
        exp_val = exp_opt

    # ④ 표준편차
    sigma_opt = st.selectbox("표준편차",
                            ['과거데이터','MLE','Vasicek','GBM','직접입력'], #'LSTM',
                            index=4, key=f"{key_prefix}_sigma_opt")
    if sigma_opt == '직접입력':
        sigma_val = st.number_input("표준편차(직접입력)",
                                    value=float(Expected_sigma.loc["Expected_sigma", sigma_key]),
                                    step=0.0001, format="%.4f", key=f"{key_prefix}_sigma_val")
    else:
        sigma_val = sigma_opt

    # ⑤ 관련자산(다중)
    related = st.multiselect("관련 자산군 선택", options, default=related_default, key=f"{key_prefix}_related")

    # ⑥ 출력 dict
    base_index_field = f"{base}_index" if use_index_suffix else base
    return {
        '기초자산_dw': base,                 # 스칼라 문자열
        '기초자산'   : base_index_field,     # suffix 여부 선택
        '모델'      : model,
        '기대수익률' : exp_val,              # '직접입력'이면 숫자, 아니면 선택 문자열
        '표준편차'   : sigma_val,             # 동일
        '관련자산'   : related,               # 리스트
        'var_dw'    : [base] + related,  # 문자열(기초+관련)
    }

def render_and_cache_asset(asset, var, T, macro, simulated_paths, params):
    paths_arr = np.asarray(simulated_paths[asset])  # (simulation_times, T)
    mean_path    = paths_arr.mean(axis=0)
    q25          = np.percentile(paths_arr, 25, axis=0)
    q50          = np.percentile(paths_arr, 50, axis=0)
    q75          = np.percentile(paths_arr, 75, axis=0)
    std_dev      = paths_arr.std(axis=0, ddof=1)

    real_index = macro.loc[macro['Date'] >= '2023-12-31', var]

    quantile_df = pd.DataFrame({'Mean': mean_path, 'Median': q50, 'Q1': q25, 'Q3': q75})
    real_df     = pd.concat([real_index]*4, axis=1); real_df.columns = ['Mean','Median','Q1','Q3']

    combined_df = pd.concat([real_df, quantile_df], ignore_index=True)
    combined_df.index = pd.date_range(start='2023-12-31', periods=len(combined_df), freq='ME')
    combined_df.index.name = 'Date'

    # --- combined_df_all 생성: real 데이터를 simulation_times만큼 복제한 뒤 simulated_paths와 concat
    simulation_times = paths_arr.shape[0]
    real_df_all = pd.concat([real_index] * simulation_times, axis=1, ignore_index=True)
    # simulated_paths를 DataFrame으로 변환
    sim_df = pd.DataFrame(paths_arr[:, 1:].T)
    # real과 simulation 결합 (열 방향으로)
    combined_df_all = pd.concat([real_df_all, sim_df], axis=0, ignore_index=True)
    combined_df_all.index = pd.date_range(start='2023-12-31', periods=len(combined_df_all), freq='ME')
    combined_df_all.index.name = 'Date'

    # --- Figure 생성 
    fig, ax = plt.subplots(figsize=(12, 6))
    sim_index = combined_df.index[-T:]
    for i in range(paths_arr.shape[0]):
        ax.plot(sim_index, paths_arr[i], color='silver', alpha=0.3, linewidth=0.6)

    ax.plot(combined_df.index, combined_df['Mean'],   label='Mean',   linewidth=2)
    ax.plot(combined_df.index, combined_df['Median'], label='Median', linestyle='--')
    ax.plot(combined_df.index, combined_df['Q1'],     label='Q1',     linestyle=':')
    ax.plot(combined_df.index, combined_df['Q3'],     label='Q3',     linestyle=':')

    ax.set_xlabel('Date'); ax.set_ylabel('Index Level')
    ax.set_title(f'{asset} - 시뮬레이션 경로')
    ax.legend(loc='best'); ax.grid(True)
    year_end_dates = combined_df.index[combined_df.index.month == 12]
    ax.set_xticks(year_end_dates); ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    fig.autofmt_xdate(); fig.tight_layout()

    # --- 세션에 저장 (덮어쓰기)
    st.session_state["plots"][asset]   = fig
    st.session_state["params"][asset]  = dict(params, **{"std_dev_mean": float(std_dev.mean())})
    st.session_state["quantile"][asset] = quantile_df
    st.session_state["real"][asset]     = real_df
    st.session_state["combined"][asset] = combined_df
    st.session_state["combined_all"][asset] = combined_df_all    
    st.session_state["simulated_paths"] = simulated_paths

# -----------------------------
# 1) 유틸 함수
# -----------------------------
def month_offset(base_dt: pd.Timestamp, liab_dt: pd.Timestamp) -> int:
    return (liab_dt.year - base_dt.year) * 12 + (liab_dt.month - base_dt.month)

def year_end_cols(W: int, o: int) -> np.ndarray:
    if o > W - 1:
        raise ValueError(f"offset {o}가 범위를 벗어남(W={W})")
    K = 1 + (W - 1 - o) // 12
    return np.array([o + 12*k for k in range(K)], dtype=int)

def build_new_ALM_DB(ALM_DB_sim: dict):
    s_keys = sorted(ALM_DB_sim.keys())
    t_keys = sorted(ALM_DB_sim[s_keys[0]].keys())
    S = len(s_keys); K = len(t_keys)
    DBO_mat = np.empty((K, S), dtype=float)
    NC_mat  = np.empty((K, S), dtype=float)
    EBP_mat = np.empty((K, S), dtype=float)
    for ti, t in enumerate(t_keys):
        for si, s in enumerate(s_keys):
            rec = ALM_DB_sim[s][t]
            DBO_mat[ti, si] = float(rec['DBO'])
            NC_mat[ti,  si] = float(rec['NC'])
            EBP_mat[ti, si] = float(rec['EBP'])
    return {'DBO': DBO_mat, 'NC': NC_mat, 'EBP': EBP_mat}

def _compute_FR(x, initial_asset_value):
    W = x.reshape(산출년수, A)
    S = A3D.shape[0]
    R_port = np.einsum('ska,ka->sk', A3D, W)  # (S, 산출년수)
    Asset = np.empty((S, 산출년수))
    FR, Surp = np.empty_like(Asset), np.empty_like(Asset)

    Asset[:, 0] = initial_asset_value
    FR[:, 0] = Asset[:, 0] / DBO_mat[0]
    Surp[:, 0] = Asset[:, 0] - DBO_mat[0]

    for t in range(1, 산출년수):
        Asset[:, t] = (Asset[:, t-1] - 0.5 * EBP_mat[t-1]) * (1 + R_port[:, t]) - 0.5 * EBP_mat[t-1] + NC_mat[t-1]
        FR[:, t] = Asset[:, t] / DBO_mat[t]
        Surp[:, t] = Asset[:, t] - DBO_mat[t]
    return R_port, Asset, FR, Surp

def _compute_FR_trimmed(x, init_val):
    R_port, Asset, FR, Surp = _compute_FR(x, init_val)
    return R_port[:, 1:], Asset[:, 1:], FR[:, 1:], Surp[:, 1:]

def make_ef_for_year(R: np.ndarray,
                    bounds: list,
                    n_target: int = 20,
                    w0: np.ndarray = None):
    S, A = R.shape
    mu   = R.mean(axis=0)              # (A,)
    Sig  = np.cov(R, rowvar=False)     # (A,A)
    mu_min, mu_max = np.quantile(mu, 0.05), np.quantile(mu, 0.95)
    if mu_min == mu_max:
        _mn, _mx = float(mu.min()), float(mu.max())
        mu_min, mu_max = (_mn, _mx) if _mx > _mn else (_mn, _mn + 0.02)
    targets = np.linspace(mu_min, mu_max, n_target)
    if w0 is None:
        w0 = np.ones(A) / A
    cons_sum1 = ({'type': 'eq', 'fun': lambda w: float(np.sum(w) - 1.0)},)
    ws, tg = [], []
    for tt in targets:
        obj  = lambda w: float(w @ Sig @ w)
        cons = cons_sum1 + ({'type': 'eq', 'fun': lambda w, ttar=tt: float(np.dot(w, mu) - ttar)},)
        res  = minimize(obj, w0, method='SLSQP', bounds=bounds, constraints=cons,
                        options={"ftol":1e-10,"maxiter":400})
        if res.success:
            ws.append(res.x); tg.append(tt)
    return np.array(ws), np.array(tg), mu, Sig, targets

def liability_year_labels(liab_dt: pd.Timestamp, o_raw: int, col_idx: np.ndarray) -> pd.DatetimeIndex:
    dates = []
    for c in col_idx:
        delta = c - o_raw
        y = liab_dt.year + (liab_dt.month - 1 + delta) // 12
        m = (liab_dt.month - 1 + delta) % 12 + 1
        dates.append(pd.Timestamp(year=y, month=m, day=1) + pd.offsets.MonthEnd(0))
    return pd.DatetimeIndex(dates)

#--------------------------------------------------------------
# Streamlit 앱 시작 

st.set_page_config(page_title="한국투자증권 ALM 시뮬레이터", page_icon=":bar_chart:",layout="wide")

st.title(" :bar_chart: 한국투자증권 ALM 시뮬레이터")

# Streamlit 앱 생성
col1, col2, col3, col4 = st.columns(4)

with col1:
    selected_date = st.date_input("기준일 입력", value = datetime(2025, 12, 31))
    st.write("기준일:", selected_date)
    기준일 = selected_date

with col2:
    산출년수 = st.number_input("산출년수 입력:", value = 6)
    st.write("산출년수:", 산출년수)

with col3:
    simulation_times = st.number_input("Simulation Time", min_value=1, value=200, step=1)    
    st.write("simulation times:", simulation_times)    

with col4:
    seed = st.number_input("SEED", min_value=1, value=999, step=1)
    st.write("SEED:", seed)    

liab_dt = pd.Timestamp(기준일)
dates = pd.date_range(
    end=liab_dt + pd.offsets.YearEnd(산출년수 - 1),
    periods=산출년수,
    freq=f"A-{liab_dt.strftime('%b').upper()}"
)
date_labels = [d.strftime("%Y-%m") for d in dates]    

# 탭 생성
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11 = st.tabs(["기초자료 업로드", "제도 설계", "제도 설계(추가)", "신입사원 가정", "금리 시나리오"\
    , "확정급여채무", "메모", "자산배분 설계", "자산배분 시뮬레이션", "자산배분", "자산배분_메모"])

# 각 탭의 내용 설정
default_url = "https://raw.githubusercontent.com/bobkim67/K-ALM/main/"
with tab1:
    #명부 업로드
    fl0 = st.file_uploader(":file_folder: 명부 Upload",type=(["csv","xlsx","xls"]))
    if fl0 is not None:
        filename = fl0.name
        명부 = pd.read_excel(fl0)
        st.info(f"✅ 사용자 업로드 파일 사용: {filename}")
    
    else:
        명부 = pd.read_excel(urljoin(default_url, quote("명부_v0.xlsx")))
        st.info("ℹ️ 업로드된 파일이 없어, GitHub 기본 파일(`명부_v0.xlsx`)을 자동 불러왔습니다.")

    명부['성별'] = 명부['식별번호'].apply(lambda x: 'M' if int(x[7]) % 2 == 1 else 'F')
    명부['dob_yr'] = 명부['식별번호'].apply(lambda x: '19' if int(x[7]) in [1, 2, 5, 6] else '20')
    명부['생년월일'] = 명부['dob_yr'] + 명부['식별번호'].str[:6]
    명부['생년월일'] = pd.to_datetime(명부['생년월일'], format='%Y%m%d').dt.strftime('%Y-%m-%d')
    명부['생년월일'] = 명부['생년월일'].apply(lambda x: datetime.strptime(x, "%Y-%m-%d").date())
    명부['생년월일'] = pd.to_datetime(명부['생년월일'], format='%Y%m%d').dt.date

    명부['입사일'] = pd.to_datetime(명부['입사일']).apply(lambda x: x.date())
    명부['기산일'] = pd.to_datetime(명부['기산일']).apply(lambda x: x.date())
    명부['기준급여'] = 명부['기준급여'].apply(lambda x: int(x))
    명부['당월퇴직추계'] = 명부['당월퇴직추계'].apply(lambda x: int(x))
    명부['연령'] = 명부['생년월일'].apply(lambda x: 나이계산(x, 기준일))
    명부['근속년수'] = (기준일 - 명부['기산일']).apply(lambda x: x.days // 365)
    명부.drop(columns=['dob_yr','식별번호'], inplace=True)
    st.dataframe(명부)
        

    #기초율 업로드
    fl1 = st.file_uploader(":file_folder: 기초율 Upload",type=(["csv","xlsx","xls"]))
    if fl1 is not None:
        filename = fl1.name
        기초율 = pd.read_excel(fl1)
        st.info(f"✅ 사용자 업로드 파일 사용: {filename}")
    else:
        기초율 = pd.read_excel(urljoin(default_url, quote("기초율_v0.xlsx")))
        st.info("ℹ️ 업로드된 파일이 없어, GitHub 기본 파일(`기초율_v0.xlsx`)을 자동 불러왔습니다.")
    st.dataframe(기초율)

    #지급률 업로드
    fl2 = st.file_uploader(":file_folder: 지급률 Upload",type=(["csv","xlsx","xls"]))
    if fl2 is not None:
        filename = fl2.name
        지급률 = pd.read_excel(fl2)
        st.info(f"✅ 사용자 업로드 파일 사용: {filename}")
    else:
        지급률 = pd.read_excel(urljoin(default_url, quote("지급률_v0.xlsx")))
        st.info("ℹ️ 업로드된 파일이 없어, GitHub 기본 파일(`지급률_v0.xlsx`)을 자동 불러왔습니다.")
    st.dataframe(지급률)

    #경제지표 업로드
    fl3 = st.file_uploader(":file_folder: 경제지표 Upload",type=(["csv","xlsx","xls"]))
    if fl3 is not None:
        filename = fl3.name
        st.write(filename)
        macro = pd.read_excel(fl3)
        st.info(f"✅ 사용자 업로드 파일 사용: {filename}")
    else:
        macro = pd.read_excel(urljoin(default_url, quote("Data_2025_v3.xlsx")))
        st.info("ℹ️ 업로드된 파일이 없어, GitHub 기본 파일(`Data_2025_v3.xlsx`)을 자동 불러왔습니다.")
    st.dataframe(macro)

    #기대수익률 업로드
    fl4 = st.file_uploader(":file_folder: 기대수익률 Upload",type=(["csv","xlsx","xls"]))
    if fl4 is not None:
        filename = fl4.name
        st.write(filename)
        Exp_rt_xl = pd.read_excel(fl4)
        st.info(f"✅ 사용자 업로드 파일 사용: {filename}")
    else:
        Exp_rt_xl = pd.read_excel(urljoin(default_url, quote("기대수익률_v0.xlsx")))
        st.info("ℹ️ 업로드된 파일이 없어, GitHub 기본 파일(`기대수익률_v0.xlsx`)을 자동 불러왔습니다.")    

    # ✅ 자산군을 인덱스로
    Exp_rt_xl.set_index("자산군", inplace=True)
    Exp_rt_xl.index.name = None

    # ✅ “기대수익률” 컬럼을 행으로 전치 (당신 코드 구조에 맞춤)
    Expected_returns = Exp_rt_xl.T
    Expected_returns.index = ["Expected_returns"]    
    st.dataframe(Expected_returns)

    cor_data = pd.DataFrame(index=macro.index)
    cor_data['Deposit_1Y'] = macro['Deposit_1Y'].diff()
    cor_data['G_Bond_1Y'] = macro['G_Bond_1Y'].diff()
    cor_data['G_Bond_3Y'] = macro['G_Bond_3Y'].diff()
    cor_data['G_Bond_5Y'] = macro['G_Bond_5Y'].diff()
    cor_data['G_Bond_10Y'] = macro['G_Bond_10Y'].diff()
    cor_data['KIS_total'] = np.log(macro['KIS_total_index']).diff()
    cor_data['KIS_IR_2Y'] = macro['KIS_IR_2Y'].diff()
    cor_data['KIS_CY_3Y'] = macro['KIS_CY_3Y'].diff()
    cor_data['KIS_IR_3Y'] = macro['KIS_IR_3Y'].diff()
    cor_data['ICE_GCI']  = macro['ICE_GCI'].diff()
    cor_data['ICE_GGI']  = macro['ICE_GGI'].diff()
    cor_data['ICE_USCI_3Y']  = macro['ICE_USCI_3Y'].diff()
    cor_data['ICE_USCI_5Y']  = macro['ICE_USCI_5Y'].diff()
    cor_data['KOSPI'] = np.log(macro['KOSPI_index']).diff()
    cor_data['MSCI_ACWI'] = np.log(macro['MSCI_ACWI_index']).diff()
    cor_data['ICE_GBMI'] = np.log(macro['ICE_GBMI_index']).diff()
    cor_data['Wage_rate'] = macro['Wage_rate']
    cor_data['Real_Estate'] = macro['Real_Estate']
    cor_data['CCI'] = macro['CCI'].diff()
  
if all([명부 is not None, 기초율 is not None, 지급률 is not None, macro is not None]):
    with tab2:
        if '직급개수' not in st.session_state:
            st.session_state.직급개수 = None

        if '제도설계' not in st.session_state:
            st.session_state.제도설계 = None
                    
        col_a, col_b = st.columns(2)
        with col_a:
            직급개수 = st.number_input("직급 개수",value=2, step=1, key="직급개수_input")
            
        #with st.form(key = '직원 설정'):
        col1, col2 = st.columns(2)
        
        with col1:
            if 직급개수 >= 1:
                st.subheader("직원 설정")
                
                직원_설계 = {
                    '직급': 1,
                    #'지급률': st.selectbox("지급률", 지급률.columns[[2,4,6,8]].tolist(), key="지급률_직원"),
                    '평가방법': st.selectbox("평가방법", ['PUC', '추계액'], key="평가방법_직원"),
                    '할당': st.selectbox("할당", ['근속년수', '지급률'], key="할당_직원"),
                    '지급률_계산' : st.selectbox("지급률 계산방식", ["일할", "월 절상", "반기 절상", "연 절상"], key = "지급률_계산_직원"),
                    '승급률': st.selectbox("승급률", 기초율.columns[5:9].tolist(), key="승급률_직원"),
                    '퇴직률': st.selectbox("퇴직률", 기초율.columns[1:5].tolist(), key="퇴직률_직원"),
                    'bu': st.number_input("임금상승률", value=0.02, key="bu_직원")
                }
                직원_설계['지급률_설정'] = st.radio("지급률 설정", ["지급률 테이블", "명부 입력"], key = "지급률설정_직원")

                if 직원_설계['지급률_설정'] == "지급률 테이블":
                    직원_설계['지급률'] = st.selectbox("지급률", 지급률.columns[[2,4,6]].tolist(), key="지급률_직원")
                        
                elif 직원_설계['지급률_설정'] == "명부 입력":
                    직원_설계['지급률'] = "명부 입력"
                    
                직원_설계['정년_설정'] = st.radio("정년 설정", ["정년 입력", "현재연령 기준"], key = "정년_설정_직원")

                if 직원_설계['정년_설정'] == "정년 입력":
                    직원_설계['정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년_직원")
                        
                elif 직원_설계['정년_설정'] == "현재연령 기준":
                    직원_설계['정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년_직원")
                    #직원_설계['정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년_직원")
                
                if 직원_설계['정년_설정'] == "현재연령 기준":
                    직원_설계['정년초과_설정'] = 직원_설계['정년_설정']
                    직원_설계['정년초과_정년'] = 직원_설계['정년']
                    직원_설계['정년초과_평가'] = 직원_설계['평가방법']
                
                else: 
                    직원_설계['정년초과_설정'] = st.radio("정년초과 설정", ["정년 입력", "현재연령 기준"], index=1, key = "정년초과_설정_직원")
                                        
                    if 직원_설계['정년초과_설정'] == "정년 입력":
                        직원_설계['정년초과_정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년초과_정년_직원")
                            
                    elif 직원_설계['정년초과_설정'] == "현재연령 기준":
                        직원_설계['정년초과_정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년초과_정년_직원")
                        #직원_설계['정년초과_정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년초과_정년_직원")        
                        
                    직원_설계['정년초과_평가'] = st.selectbox("정년초과 평가", ['PUC', '추계액'], index=1, key="정년초과_평가_직원")       
            else:
                st.write("직원 설정을 활성화하려면 직급 개수를 증가시켜주세요.")
        
        with col2:
            # 임원 설정
            if 직급개수 >= 2:
                st.subheader("임원 설정")
                임원_설계 = {
                    '직급': 2,
                    #'지급률': st.selectbox("지급률", 지급률.columns[[2,4,6,8]].tolist(), key="지급률_임원"),
                    '평가방법': st.selectbox("평가방법", ['PUC', '추계액'], index=1, key="평가방법_임원"),
                    '할당': st.selectbox("할당", ['근속년수', '지급률'], key="할당_임원"),
                    '지급률_계산' : st.selectbox("지급률 계산방식", ["일할", "월 절상", "반기 절상", "연 절상"], key = "지급률_계산_임원"),
                    '승급률': st.selectbox("승급률", 기초율.columns[5:9].tolist(), key="승급률_임원"),
                    '퇴직률': st.selectbox("퇴직률", 기초율.columns[1:5].tolist(), key="퇴직률_임원"),
                    'bu': st.number_input("임금상승률", value=0.02, key="bu_임원")
                }
                임원_설계['지급률_설정'] = st.radio("지급률 설정", ["지급률 테이블", "명부 입력"], key = "지급률설정_임원")

                if 임원_설계['지급률_설정'] == "지급률 테이블":
                    임원_설계['지급률'] = st.selectbox("지급률", 지급률.columns[[2,4,6]].tolist(), key="지급률_임원")
                        
                elif 임원_설계['지급률_설정'] == "명부 입력":
                    임원_설계['지급률'] = "명부 입력"

                임원_설계['정년_설정'] = st.radio("정년 설정", ["정년 입력", "현재연령 기준"], index=1, key = "정년_설정_임원")
                
                if 임원_설계['정년_설정'] == "정년 입력":
                    임원_설계['정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년_임원")
                        
                elif 임원_설계['정년_설정'] == "현재연령 기준":
                    임원_설계['정년'] = st.number_input("현재연령 기준", value=6, step=1, key="정년_임원")
                    #임원_설계['정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년_임원")

                if 임원_설계['정년_설정'] == "현재연령 기준":
                    임원_설계['정년초과_설정'] = 임원_설계['정년_설정']
                    임원_설계['정년초과_정년'] = 임원_설계['정년']
                    임원_설계['정년초과_평가'] = 임원_설계['평가방법']            
                
                else:
                    임원_설계['정년초과_설정'] = st.radio("정년초과 설정", ["정년 입력", "현재연령 기준"], key = "정년초과_설정_임원")
                    
                    if 임원_설계['정년초과_설정'] == "정년 입력":
                        임원_설계['정년초과_정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년초과_정년_임원")
                            
                    elif 임원_설계['정년초과_설정'] == "현재연령 기준":
                        임원_설계['정년초과_정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년초과_정년_임원")
                        #임원_설계['정년초과_정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년초과_정년_임원")        
                    
                    임원_설계['정년초과_평가'] = st.selectbox("정년초과 평가", ['PUC', '추계액'], index=1, key="정년초과_평가_임원")       
            else:
                st.write("임원 설정을 활성화하려면 직급 개수를 증가시켜주세요.")

        if st.button(label="저장", key = '제도설정'):
            st.session_state.직급개수 = 직급개수
            if st.session_state.직급개수 == 1:
                st.session_state.제도설계 = pd.DataFrame([직원_설계], index = ['직원'])
                st.write("직원 설정:", st.session_state.제도설계)
                    
            elif st.session_state.직급개수 >= 2:
                st.session_state.제도설계 = pd.concat([pd.DataFrame([직원_설계], index = ['직원']), pd.DataFrame([임원_설계], index = ['임원'])])
                st.write("직원 설정:", st.session_state.제도설계)
            
    with tab3:
        if '직급개수2' not in st.session_state:
            st.session_state.직급개수2 = None

        if '제도설계' not in st.session_state:
            st.session_state.제도설계 = None 
            
        col_a, col_b = st.columns(2)
        with col_a:
            직급개수2 = st.number_input("직급 개수",value=0, step=1, key="직급개수2_input")

    # 직원 설정      
        col1, col2 = st.columns(2)

        with col1:
            if 직급개수2 >= 1:
                st.subheader("직원1 설정")
                직원1_설계 = {
                    '직급': 3,
                    #'지급률': st.selectbox("지급률", 지급률.columns[[2,4,6,8]].tolist(), key="지급률_직원1"),
                    '평가방법': st.selectbox("평가방법", ['PUC', '추계액'], key="평가방법_직원1"),
                    '할당': st.selectbox("할당", ['근속년수', '지급률'], key="할당_직원1"),
                    '지급률_계산': st.selectbox("지급률 계산방식", ["일할", "월 절상", "반기 절상", "연 절상"], key = "지급률_계산_직원1"),
                    '승급률': st.selectbox("승급률", 기초율.columns[5:9].tolist(), key="승급률_직원1"),
                    '퇴직률': st.selectbox("퇴직률", 기초율.columns[1:5].tolist(), key="퇴직률_직원1"),
                    'bu': st.number_input("임금상승률", value=0.02, key="bu_직원1")
                }
                직원1_설계['지급률_설정'] = st.radio("지급률 설정", ["지급률 테이블", "명부 입력"], key = "지급률설정_직원1")

                if 직원1_설계['지급률_설정'] == "지급률 테이블":
                    직원1_설계['지급률'] = st.selectbox("지급률", 지급률.columns[[2,4,6]].tolist(), key="지급률_직원1")
                        
                elif 직원1_설계['지급률_설정'] == "명부 입력":
                    직원1_설계['지급률'] = "명부 입력"

                직원1_설계['정년_설정'] = st.radio("정년 설정", ["정년 입력", "현재연령 기준"], key = "정년_설정_직원1")
                
                if 직원1_설계['정년_설정'] == "정년 입력":
                    직원1_설계['정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년_직원1")
                        
                elif 직원1_설계['정년_설정'] == "현재연령 기준":
                    직원1_설계['정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년_직원1")
                    #직원1_설계['정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년_직원1")

                if 직원1_설계['정년_설정'] == "현재연령 기준":
                    직원1_설계['정년초과_설정'] = 직원1_설계['정년_설정']
                    직원1_설계['정년초과_정년'] = 직원1_설계['정년']
                    직원1_설계['정년초과_평가'] = 직원1_설계['평가방법']            
                
                else:                
                    직원1_설계['정년초과_설정'] = st.radio("정년초과 설정", ["정년 입력", "현재연령 기준"], index=1, key = "정년초과_설정_직원1")
                    
                    if 직원1_설계['정년초과_설정'] == "정년 입력":
                        직원1_설계['정년초과_정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년초과_정년_직원1")
                            
                    elif 직원1_설계['정년초과_설정'] == "현재연령 기준":
                        직원1_설계['정년초과_정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년초과_정년_직원1")
                        #직원1_설계['정년초과_정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년초과_정년_직원1")        
                    
                    직원1_설계['정년초과_평가'] = st.selectbox("정년초과 평가", ['PUC', '추계액'], index=1, key="정년초과_평가_직원1")       
                
            else:
                st.write("직원1 설정을 활성화하려면 직급 개수를 증가시켜주세요.")

        with col2:
            if 직급개수2 >= 2:
                st.subheader("직원2 설정")
                직원2_설계 = {
                    '직급': 4,
                    #'지급률': st.selectbox("지급률", 지급률.columns[[2,4,6,8]].tolist(), key="지급률_직원2"),
                    '평가방법': st.selectbox("평가방법", ['PUC', '추계액'], key="평가방법_직원2"),
                    '할당': st.selectbox("할당", ['근속년수', '지급률'], key="할당_직원2"),
                    '지급률_계산': st.selectbox("지급률 계산방식", ["일할", "월 절상", "반기 절상", "연 절상"], key = "지급률_계산_직원2"),
                    '승급률': st.selectbox("승급률", 기초율.columns[5:9].tolist(), key="승급률_직원2"),
                    '퇴직률': st.selectbox("퇴직률", 기초율.columns[1:5].tolist(), key="퇴직률_직원2"),
                    'bu': st.number_input("임금상승률", value=0.02, key="bu_직원2")
                }
                직원2_설계['지급률_설정'] = st.radio("지급률 설정", ["지급률 테이블", "명부 입력"], key = "지급률설정_직원2")

                if 직원2_설계['지급률_설정'] == "지급률 테이블":
                    직원2_설계['지급률'] = st.selectbox("지급률", 지급률.columns[[2,4,6]].tolist(), key="지급률_직원2")
                        
                elif 직원2_설계['지급률_설정'] == "명부 입력":
                    직원2_설계['지급률'] = "명부 입력"

                직원2_설계['정년_설정'] = st.radio("정년 설정", ["정년 입력", "현재연령 기준"], key = "정년_설정_직원2")
                
                if 직원2_설계['정년_설정'] == "정년 입력":
                    직원2_설계['정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년_직원2")
                        
                elif 직원2_설계['정년_설정'] == "현재연령 기준":
                    직원2_설계['정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년_직원2")
                    #직원2_설계['정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년_직원2")

                if 직원2_설계['정년_설정'] == "현재연령 기준":
                    직원2_설계['정년초과_설정'] = 직원2_설계['정년_설정']
                    직원2_설계['정년초과_정년'] = 직원2_설계['정년']
                    직원2_설계['정년초과_평가'] = 직원2_설계['평가방법']            
                
                else:                  
                    직원2_설계['정년초과_설정'] = st.radio("정년초과 설정", ["정년 입력", "현재연령 기준"], index=1, key = "정년초과_설정_직원2")
                    
                    if 직원2_설계['정년초과_설정'] == "정년 입력":
                        직원2_설계['정년초과_정년'] = st.number_input("정년 (정수 입력)", value=60, step=1, key="정년초과_정년_직원2")
                            
                    elif 직원2_설계['정년초과_설정'] == "현재연령 기준":
                        직원2_설계['정년초과_정년'] = st.number_input("현재연령 기준", value=1, step=1, key="정년초과_정년_직원2")
                        #직원2_설계['정년초과_정년'] = st.selectbox("현재연령 기준", ['+1', '+2', '+3', '+4', '+5', '+6', '+7', '+8', '+9', '+10'], key="정년초과_정년_직원2")        
                    
                    직원2_설계['정년초과_평가'] = st.selectbox("정년초과 평가", ['PUC', '추계액'], index=1, key="정년초과_평가_직원2")       
                
            else:
                st.write("직원2 설정을 활성화하려면 직급 개수를 증가시켜주세요.")
        
        if st.button(label="저장", key = '제도설정_추가'):
            st.session_state.직급개수2 = 직급개수2
            if st.session_state.직급개수 == 1:
                if st.session_state.직급개수2 == 1:
                    st.session_state.제도설계 = pd.concat([pd.DataFrame([직원_설계], index = ['직원'])\
                        ,pd.DataFrame([직원1_설계], index = ['직원1'])])
                    st.write("직원 설정:", st.session_state.제도설계)
                    
                elif st.session_state.직급개수2 == 2:
                    st.session_state.제도설계 = pd.concat([pd.DataFrame([직원_설계], index = ['직원'])\
                        , pd.DataFrame([직원1_설계], index = ['직원1']), pd.DataFrame([직원2_설계], index = ['직원2'])])
                    st.write("직원 설정:", st.session_state.제도설계)
            
            if st.session_state.직급개수 == 2:
                if st.session_state.직급개수2 == 1:
                    st.session_state.제도설계 = pd.concat([pd.DataFrame([직원_설계], index = ['직원']), pd.DataFrame([임원_설계], index = ['임원'])\
                        ,pd.DataFrame([직원1_설계], index = ['직원1'])])
                    st.write("직원 설정:", st.session_state.제도설계)
                    
                elif 직급개수2 == 2:
                    st.session_state.제도설계 = pd.concat([pd.DataFrame([직원_설계], index = ['직원']), pd.DataFrame([임원_설계], index = ['임원'])\
                        , pd.DataFrame([직원1_설계], index = ['직원1']), pd.DataFrame([직원2_설계], index = ['직원2'])])
                    st.write("직원 설정:", st.session_state.제도설계)                   

    with tab4:
        # 앱의 시작 부분에 세션 상태 초기화
        if '신입명부' not in st.session_state:
            st.session_state.신입명부 = pd.DataFrame()
                
        col_a, col_b = st.columns(2)
        with col_a:
            신입가정 = st.number_input("신입가정",value=0, step=1, key="신입가정")
    
        col1, col2 = st.columns(2)
        
        with col1:
            if 신입가정 >= 1:
                st.subheader("신입가정 설정")
                신입가정1 = {
                    '직급': st.number_input("직급",value=1, step=1, key="직급_신입1"),
                    '성별': st.selectbox("성별", ['M', 'F'], key="성별_신입1"),
                    '연령': st.number_input("연령",value=30, step=1, key="연령_신입1"),
                    '입사일': st.selectbox("입사일", ['연초', '연중', '연말'], index = 1, key="입사일_신입1"),
                    '기준급여': st.number_input("기준급여",value=3000000, step=1000, key="급여_신입1"),
                    '비중': st.number_input("비중", min_value=0.0, max_value=1.0, value=1.0, step=0.01, key="비중_신입1"),
                }
                
            else:
                st.write("신입가정1 설정을 활성화하려면 직급 개수를 증가시켜주세요.")

        with col2:
            if 신입가정 >= 2:
                st.subheader("신입가정 설정")
                신입가정2 = {
                    '직급': st.number_input("직급",value=1, step=1, key="직급_신입2"),
                    '성별': st.selectbox("성별", ['M', 'F'], index = 1, key="성별_신입2"),
                    '연령': st.number_input("연령",value=30, step=1, key="연령_신입2"),
                    '입사일': st.selectbox("입사일", ['연초', '연중', '연말'], index = 1, key="입사일_신입2"),
                    '기준급여': st.number_input("기준급여",value=3000000, step=1000, key="급여_신입2"),
                    '비중': st.number_input("비중", min_value=0.0, max_value=1.0, value=0.5, step=0.01, key="비중_신입2"),
                }

            else:
                st.write("신입가정2 설정을 활성화하려면 직급 개수를 증가시켜주세요.")

        if st.button(label="저장", key = '제도설정_신규'):
            if 신입가정 == 1:
                if 신입가정1['비중'] != 1:
                    st.write("가정 별 비중 합계 값이 100%가 되도록 재설정")
                else:
                    신입명부 = pd.DataFrame([신입가정1], index = [1])
                    신입명부.loc[신입명부['입사일'] == '연중', '입사일_date'] = date(기준일.year, 기준일.month, 1) + relativedelta(months=+7)
                    신입명부.loc[신입명부['입사일'] == '연초', '입사일_date'] = date(기준일.year + 1, 기준일.month, 1)
                    신입명부.loc[신입명부['입사일'] == '연말', '입사일_date'] = date(기준일.year, 기준일.month, 기준일.day) + relativedelta(months=+12)
                    st.session_state.신입명부 = 신입명부
                    st.write("직원 설정:", 신입명부.iloc[:, :-1])
                    
            elif 신입가정 >= 2:
                if 신입가정1['비중'] + 신입가정2['비중'] != 1:
                    st.write("가정 별 비중 합계 값이 100%가 되도록 재설정")
                else:
                    신입명부 = pd.concat([pd.DataFrame([신입가정1], index = [1]), pd.DataFrame([신입가정2], index = [2])])
                    신입명부.loc[신입명부['입사일'] == '연중', '입사일_date'] = date(기준일.year, 기준일.month, 1) + relativedelta(months=+7)
                    신입명부.loc[신입명부['입사일'] == '연초', '입사일_date'] = date(기준일.year + 1, 기준일.month, 1)
                    신입명부.loc[신입명부['입사일'] == '연말', '입사일_date'] = date(기준일.year, 기준일.month, 기준일.day) + relativedelta(months=+12)
                    st.session_state.신입명부 = 신입명부
                    st.write("직원 설정:", 신입명부.iloc[:, :-1])

    with tab5:
        projection_period = 산출년수 
        T  = 산출년수 * 12
        dt = 1/12
        index_rows = range(simulation_times)
        columns_ts = range(1, T + 1)

        st.subheader('시나리오 Simulation')
        # ▶ 선택 UI
        var_asset = st.multiselect(
            "자산군 선택",
            ['G_Bond_10Y','KIS_total','KOSPI','MSCI_ACWI','Real_Estate','Deposit_1Y','Wage_rate'],
            default=['G_Bond_10Y']
        )
        var_related = st.multiselect(
            "관련 자산군 선택",
            macro.columns[1:],
            default=['KIS_total','KOSPI','MSCI_ACWI','Real_Estate','Deposit_1Y','Wage_rate']
        )

        # ▶ 설계 / 매핑: 먼저 정의 (아래 Expected_*에서 참조)
        asset_related = {'Interest_rate': var_asset + var_related}
        var_dw = sorted(set(v for varlist in asset_related.values() for v in varlist))
        variable_number = len(var_dw)

        # ▶ Expected Returns (연율 → 로그수익률)
        np.random.seed(seed)

        # Expected_returns.loc["Expected_returns", "Domestic_bond"]  = 0.0435
        # Expected_returns.loc["Expected_returns", "Global_bond"]    = 0.0423
        # Expected_returns.loc["Expected_returns", "Domestic_stock"] = 0.0630
        # Expected_returns.loc["Expected_returns", "Global_stock"]   = 0.0763
        # Expected_returns.loc["Expected_returns", "PIGP"]           = 0.0310
        Expected_returns.loc["Expected_returns", "Interest_rate"]  = vasicek_calibrate(macro[asset_related["Interest_rate"][0]], dt=dt)[1]
        # Expected_returns.loc["Expected_returns", "Real_Estate"]    = 0.0630

        for a in ["Domestic_bond","Domestic_stock","Global_stock","Global_bond"]:
            Expected_returns.loc["Expected_returns", a] = np.log(1 + Expected_returns.at['Expected_returns', a])

        Expected_sigma = pd.DataFrame()
        # ▶ Expected Sigma
        Expected_sigma.loc["Expected_sigma", "Domestic_bond"]  = np.log(macro['KIS_total_index']/macro['KIS_total_index'].shift(1)).dropna().std(ddof=1) * np.sqrt(12) * 2
        Expected_sigma.loc["Expected_sigma", "Domestic_stock"] = np.log(macro['KOSPI_index']/macro['KOSPI_index'].shift(1)).dropna().std(ddof=1) * np.sqrt(12) * 2
        Expected_sigma.loc["Expected_sigma", "Global_bond"]   = np.log(macro['ICE_GBMI_index']/macro['ICE_GBMI_index'].shift(1)).dropna().std(ddof=1) * np.sqrt(12) * 2
        Expected_sigma.loc["Expected_sigma", "Global_stock"]    = np.log(macro['MSCI_ACWI_index']/macro['MSCI_ACWI_index'].shift(1)).dropna().std(ddof=1) * np.sqrt(12) * 2
        Expected_sigma.loc["Expected_sigma", "PIGP"] = macro['Deposit_1Y'].std(ddof=1) * np.sqrt(12)
        Expected_sigma.loc["Expected_sigma", "Interest_rate"] = macro['G_Bond_10Y'].std(ddof=1) * np.sqrt(12)
        Expected_sigma.loc["Expected_sigma", "Real_Estate"] = macro['Real_Estate'].std(ddof=1) * np.sqrt(12)


        # ▶ 사용자 직접 입력(기본값: Interest_rate에서 가져옴)
        자산군_모델 = st.selectbox("시뮬레이션 모델", ['Vasicek','GBM'], index=0, key="설계_모델") # 'LSTM'

        기대수익률_value = st.number_input(
            "기대수익률(직접입력)",
            value=float(Expected_returns.loc["Expected_returns", "Interest_rate"]),
            step=0.0001,
            format="%.4f",
            key="기대수익률_값"
        )
        표준편차_value = st.number_input(
            "표준편차(직접입력)",
            value=float(Expected_sigma.loc["Expected_sigma", "Interest_rate"]),
            step=0.0001,
            format="%.4f",
            key="표준편차_값"
        )
        
        자산군_설계 = {
            '기초자산_dw': var_asset,
            '기초자산': var_asset,
            '모델': 자산군_모델,
            '기대수익률': 기대수익률_value,
            '표준편차': 표준편차_value,
            '관련자산': var_related
        }

        # 선택 재반영(원하면 아래 한 줄로 교체 가능)
        asset_related = {'Interest_rate': 자산군_설계['기초자산_dw'] + 자산군_설계['관련자산']}
        var_dw = sorted(set(v for varlist in asset_related.values() for v in varlist))
        variable_number = len(var_dw)

        # ▶ 시뮬레이션 실행
        data, ci, combined_i, moving_avg = {}, {}, {}, {}
        simulated_paths_i, kappa_dict = {}, {}

        if st.button(label="실행", key='vasicek_simul'):
            # ▶ 상관행렬 & Cholesky (보정 없음)
            correlation_df = cor_data[var_dw].corr()
            correlation = correlation_df.values
            L = np.linalg.cholesky(correlation)

            # ▶ epsilon 생성
            all_epsilon = {asset: pd.DataFrame(index=index_rows, columns=columns_ts, dtype=float) for asset in asset_related}
            for x in range(1, T + 1):
                Z = np.random.normal(0, 1, size=(variable_number, simulation_times))
                correlated = L @ Z
                epsilon_df = pd.DataFrame(correlated.T, columns=var_dw)
                for asset, vars_list in asset_related.items():
                    valid_vars = [v for v in vars_list if v in epsilon_df.columns]
                    all_epsilon[asset][x] = epsilon_df[valid_vars].mean(axis=1).values            
            col1, col2 = st.columns(2)
            with col1:
                sns.set(style='white'); cmap = sns.diverging_palette(220, 20, as_cmap=True)
                sns.heatmap(correlation_df, annot=True, cmap=cmap, center=0, fmt=".2f",
                            square=True, linewidths=.5, cbar_kws={"shrink": 0.8})
                plt.title('Correlation Heatmap')
                st.pyplot(plt); plt.savefig('Correlation_Heatmap.png'); plt.clf()

            with col2:
                sns.set(style='white'); cmap = sns.diverging_palette(220, 20, as_cmap=True)
                sns.heatmap(pd.DataFrame(L, index=var_dw, columns=var_dw), annot=True, cmap=cmap, center=0, fmt=".2f",
                            square=True, linewidths=.5, cbar_kws={"shrink": 0.8})
                plt.title('Cholesky Factor Heatmap')
                st.pyplot(plt); plt.savefig('Correlation(Cholesky)_Heatmap.png'); plt.clf()

            # ── 자산별 시뮬
            for asset in list(asset_related.keys()):
                var = 자산군_설계['기초자산_dw']                
                sim_type = 자산군_모델
                r_expected  = 기대수익률_value
                sd_expected = 표준편차_value
                eps = all_epsilon[asset].values  # (sim_times, T)

                if sim_type == 'Vasicek':
                    rates = macro[var].dropna()
                    kappa0, theta0, sigma0, r0 = vasicek_calibrate(rates, dt=dt)
                    theta = theta0
                    sigma = sd_expected
                    target_level = theta

                    res = minimize(
                        vasicek_objective,
                        x0=max(kappa0, 1e-3),
                        args=(theta, sigma, r0, simulation_times, T, dt, eps, target_level),
                        method="SLSQP",
                        bounds=[(0.1, 0.5)]
                    )
                    kappa = float(res.x[0]); kappa_dict[asset] = kappa
                    paths = vasicek_simul(r0, kappa, theta, sigma, simulation_times, T, dt, eps)
                    simulated_paths_i[asset] = paths  # (sim_times, T)

            # ── 경로 플롯
            col3, col4 = st.columns(2)
        
            with col3:
                plt.figure(); tgrid = np.linspace(0, projection_period, T)
                for i in range(simulation_times): plt.plot(tgrid, simulated_paths_i[asset][i])
                plt.xlabel('Year'); plt.ylabel(f'{asset}')
                plt.title(f'{asset} Simulations'); plt.grid(True)
                st.pyplot(plt); plt.savefig(f'{asset}_Vasicek.png'); plt.clf()
            
            col5, col6 = st.columns(2)
            with col5:
                # ── 요약/CI/결합 시계열
                # 2) 실측 시계열 준비 (실측 인덱스 그대로)
                hist = macro[var].dropna()
                hist_index = pd.to_datetime(macro['Date']).iloc[-len(hist):]  # 해당 열 길이에 맞춰 정렬
                hist.index = hist_index
            
                # 3) 시뮬 결과 확인/요약 계산
                paths = simulated_paths_i.get(asset)
                arr = np.asarray(paths)  # (simulation_times, T)

                # 미래 인덱스: 실측 마지막 달 '다음 달'부터 T개월
                last_hist = hist.index[-1]
                future_index = pd.date_range(start=last_hist + pd.offsets.MonthEnd(1), periods=T, freq='M')

                # 요약선 계산 (time-axis: axis=0)
                min_line   = np.min(arr, axis=0)
                q1_line    = np.percentile(arr, 25, axis=0)
                med_line   = np.percentile(arr, 50, axis=0)
                q3_line    = np.percentile(arr, 75, axis=0)
                max_line   = np.max(arr, axis=0)
                mean_line  = np.mean(arr, axis=0)
                # 95% CI (정규 근사 or 이미 계산된 값이 있으면 그걸 사용)
                std_line   = np.std(arr, axis=0, ddof=0)
                z_score    = 1.96
                ci_low     = mean_line - z_score * std_line / np.sqrt(arr.shape[0])
                ci_high    = mean_line + z_score * std_line / np.sqrt(arr.shape[0])

                # 4) 플롯: 실측 + 개별 경로(연한 회색) + 요약선(컬러)
                fig, ax = plt.subplots(figsize=(10, 5))

                # (a) 실측
                ax.plot(hist.index, hist.values, label="Actual", linewidth=1.8)

                # (b) 개별 경로 (연한 회색)
                for i in range(arr.shape[0]):
                    ax.plot(future_index, arr[i], color=(0.7, 0.7, 0.7), alpha=0.3, linewidth=0.8)

                # (c) 요약선 (색상 구분)
                ax.plot(future_index, min_line,  label="Min",    linewidth=1.2)
                ax.plot(future_index, q1_line,   label="Q1",     linewidth=1.2)
                ax.plot(future_index, med_line,  label="Median", linewidth=1.8)
                ax.plot(future_index, q3_line,   label="Q3",     linewidth=1.2)
                ax.plot(future_index, max_line,  label="Max",    linewidth=1.2)
                ax.plot(future_index, mean_line, label="Mean",   linewidth=1.6)

                # (d) 95% CI는 점선으로
                ax.plot(future_index, ci_low,  label="95% CI-", linestyle="--", linewidth=1.0)
                ax.plot(future_index, ci_high, label="95% CI+", linestyle="--", linewidth=1.0)

                # 18개월 간격, 연말(anchor=12/1)에서 시작 → 월말로 이동
                ticks_18m = pd.date_range(
                    start=pd.Timestamp(pd.Timestamp(hist.index.min().date()).year, 12, 1),
                    end=pd.Timestamp(max(hist.index.max(), future_index.max()).date()),
                    freq="18MS"                   # 18-Month Start anchored at December
                ) + pd.offsets.MonthEnd(0)     # 12/31 -> (18개월 뒤) 6/30 -> (18개월 뒤) 12/31 ...

                ax.set_xticks(ticks_18m)
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
                fig.autofmt_xdate()

                ax.set_title(f"{asset}: Actual + Simulated ({T/12:.0f}y)")
                ax.set_xlabel("Date")
                ax.set_ylabel(asset)
                ax.grid(True)
                st.pyplot(fig)

                combined_df = pd.DataFrame(index=hist.index.append(future_index))
                combined_df.index.name = "Date"

                # 과거: Actual만 채움
                for col in ["Actual","Min","Q1","Median","Q3","Max","Mean","95% CI-","95% CI+"]:
                    combined_df.loc[hist.index, col] = hist.values

                # 미래: 요약선 채움
                future_df = pd.DataFrame(
                    {
                        "Min":     min_line,
                        "Q1":      q1_line,
                        "Median":  med_line,
                        "Q3":      q3_line,
                        "Max":     max_line,
                        "Mean":    mean_line,
                        "95% CI-": ci_low,
                        "95% CI+": ci_high,
                    },
                    index=future_index
                )
                combined_df = combined_df.combine_first(future_df)
                st.write(f"{asset} preview:")
                st.dataframe(combined_df.sort_index(ascending=False))
                st.session_state.combined_i = combined_df             

            with col6:
                # 0) 입력 정리
                ma_input = st.session_state.get('combined_i', {}).copy()
                ma_input.index = pd.to_datetime(ma_input.index)             # 인덱스 Datetime 보장
                if 'Actual' in ma_input.columns and 'Mean' in ma_input.columns:
                    ma_input['Actual'] = ma_input['Actual'].fillna(ma_input['Mean'])

                # 1) 36M 이동평균
                ma = ma_input.rolling(window=36, min_periods=36).mean()

                # 2) 유효 구간 (모든 컬럼 NaN인 행 제거 후 범위 산정)
                ma_valid = ma.dropna(how="all")
                first_valid_dt = ma_valid.index.min()
                max_dt         = ma.index.max()

                # 3) 플롯: “Actual+Simulated” 스타일과 동일하게 ax.plot로만 그리기
                fig, ax = plt.subplots(figsize=(10, 5))

                # (a) 컬럼별로 직접 그리기
                for col in ma.columns:
                    s = ma[col]
                    if s.notna().any():
                        ax.plot(s.index, s.values, label=col, linewidth=1.4)

                # (b) 18개월 간격, 연말(anchor=12/1)에서 시작 → 월말로 이동 (너 코드와 동일한 로직)
                import matplotlib.dates as mdates
                anchor_start = pd.Timestamp(first_valid_dt.year, 12, 1)     # 12/1에서 시작
                ticks_18m = pd.date_range(
                    start=anchor_start,
                    end=max_dt,
                    freq="18MS",             # 18-Month Start anchored at December
                ) + pd.offsets.MonthEnd(0)   # 월말(12/31, 6/30, ...)로 이동
                # 범위 내 필터
                ticks_18m = [t for t in ticks_18m if (t >= first_valid_dt) and (t <= max_dt)]

                # (c) 축/포맷
                ax.set_xlim(first_valid_dt, max_dt)
                ax.set_xticks(ticks_18m)
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
                fig.autofmt_xdate()

                ax.set_title("36M Moving Average")
                ax.set_xlabel("Date")
                ax.grid(True)
                ax.legend(ncol=min(4, len(ma.columns)))   # 컬럼 많을 가능성 대비
                st.pyplot(fig)

                st.write("36M Moving Average preview:")
                st.dataframe(ma.sort_index(ascending=False))
                st.session_state.moving_avg = ma

                # 4) 할인율 DB용: 0.25% 단위 반올림(단, 마이너스 금리는 배제(0으로 고정))
                rates_db = ma.copy()
                rates_round = pd.DataFrame(index=rates_db.index)
                cols = ['Min', 'Q1', 'Median', 'Q3', 'Max', '95% CI-', '95% CI+', 'Mean']
                for col in cols:
                    rates_round[col] = (
                        (rates_db[col] / 0.0025).round().clip(lower=0) * 0.0025
                    ).round(4)
                기준일_dt = pd.to_datetime(기준일)
                st.dataframe(rates_round.sort_index(ascending=False))
                st.session_state.rates_round = rates_round

            combined_i = st.session_state.get('combined_i', {})
            moving_avg = st.session_state.get('moving_avg', {})
            
            # ▶ 엑셀 파일 생성
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                combined_i.to_excel(writer, sheet_name="RAW", index=True)
                moving_avg.to_excel(writer, sheet_name="MA", index=True)
            output.seek(0)

            st.download_button(
                label="Projection & Moving Average 내려받기",
                data=output.getvalue(),
                file_name="Rates.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with tab6:
        if 'ALM_DB' not in st.session_state or 'ALM_DB_ind' not in st.session_state or 'rates_round_filtered' not in st.session_state:
            # 초기화 및 데이터 저장
            st.session_state.ALM_DB = {}
            st.session_state.ALM_DB_ind = {}
            st.session_state.cf = {}
            st.session_state.할인율 = {}
            st.session_state.명부_dict = {}
            st.session_state.rates_round_filtered = {}
        if 'DBO_plot' not in st.session_state:
            st.session_state.DBO_plot = None
        if 'NC_plot' not in st.session_state:
            st.session_state.NC_plot = None
        if 'EBP_plot' not in st.session_state:
            st.session_state.EBP_plot = None
            
        st.write("제도설계")
        st.write(st.session_state.제도설계)
        st.write("신입가정")
        st.write(st.session_state.신입명부.iloc[:, :-1])
        ALM_DB = {}
        ALM_DB_ind = {}
        cf = {}
        results_dict = {}
        명부_dict = {}
        할인율 = []
        retired = []
        if st.button(label="퇴직부채 산출", key = 'DBO_simul'):
            # 할인율 0.25% 단위 반올림(단, 마이너스 금리는 배제(0으로 고정))
            rates_db = st.session_state.moving_avg.drop(columns=['Actual'], errors='ignore')
            cols = ['Min', 'Q1', 'Median', 'Q3', 'Max', '95% CI-', '95% CI+', 'Mean']
            for col in cols:
                rates_db[col] = (
                    (rates_db[col] / 0.0025).round().clip(lower=0) * 0.0025
                ).round(4)
            기준일_dt = pd.to_datetime(기준일)
            rates_filtered = rates_db[rates_db.index >= 기준일_dt]

            # 기준일과 같은 "월-일"만 추출
            월, 일 = 기준일_dt.month, 기준일_dt.day 
            rates_round_filtered = rates_filtered[
                (rates_filtered.index.month == 월) & (rates_filtered.index.day == 일)
            ]
            strt = rates_db.index.get_loc(기준일_dt) + 1

            for j in range(0, len(rates_round_filtered.columns)):
                interval = rates_round_filtered.columns[j]
                ALM_DB[interval] = {}
                ALM_DB_ind[interval] = {}
                results_dict[interval] = {}
                cf[interval] = {} 
            
                for sim_yr in range(0, 산출년수): 
                    시산기준일 = date(기준일.year + sim_yr, 기준일.month, 기준일.day)
                    할인율 = rates_round_filtered.iloc[sim_yr, j]
                    result_table, 명부, cf_data = simul(명부, 기준일, sim_yr, 할인율, 0)
                    명부_dict[sim_yr] = 명부
                    total_emp = len(명부)
                    print(f"시산기준일: {시산기준일}, Discount Rate: {할인율}, interval: {interval}, Year_{sim_yr}: Number of employees: {total_emp}")
                    
                    # Save the result_table to the results_dict with key 'Year_0' or 'Year_1'
                    results_dict[interval][f"Year_{sim_yr}"] = result_table
                    ALM_DB[interval][sim_yr] = {
                        'DBO' : results_dict[interval][f"Year_{sim_yr}"]['DBO'].sum(),
                        'NC' : results_dict[interval][f"Year_{sim_yr}"]['NC'].sum(),
                        'EBP' : results_dict[interval][f"Year_{sim_yr}"]['EBP'].sum(),
                    }
                    ALM_DB_ind[interval][sim_yr] = results_dict[interval][f"Year_{sim_yr}"]
                    cf[interval][f"Year_{sim_yr}"] = cf_data

                명부 = 명부_dict[0]
        
            st.session_state.ALM_DB = ALM_DB
            st.session_state.ALM_DB_ind = ALM_DB_ind
            st.session_state.명부_dict = 명부_dict
            st.session_state.cf = cf
            st.session_state.rates_round_filtered = rates_round_filtered

        
        if "rates_round_filtered" in st.session_state:
            st.write("할인율(국고채 10년 36MMA)")
            st.write(st.session_state.rates_round_filtered)     
        df_DBO = pd.DataFrame({col: {row: value['DBO'] for row, value in st.session_state.ALM_DB[col].items()} for col in st.session_state.ALM_DB.keys()})
        df_NC = pd.DataFrame({col: {row: value['NC'] for row, value in st.session_state.ALM_DB[col].items()} for col in st.session_state.ALM_DB.keys()})
        df_EBP = pd.DataFrame({col: {row: value['EBP'] for row, value in st.session_state.ALM_DB[col].items()} for col in st.session_state.ALM_DB.keys()})
        
        # 다운로드 버튼을 누를 준비가 되었는지 확인하는 플래그
        generate_excel = False

        # '엑셀 파일 생성' 버튼 추가
        if st.button(label="엑셀 파일 생성", key = "DBO_download"):
            generate_excel = True

        # 엑셀 파일 생성 
        if generate_excel:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df_DBO.to_excel(writer, sheet_name='DBO', index=True)
                df_NC.to_excel(writer, sheet_name='NC', index=True)
                df_EBP.to_excel(writer, sheet_name='EBP', index=True)
            excel_data = output.getvalue()

            # 다운로드 버튼 생성
            st.download_button(
                label="데이터 내려받기",
                data=excel_data,
                file_name="ALM_DBO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
                
        col1, col2 = st.columns(2)   
        with col1:
            df_DBO = pd.DataFrame({col: {row: value['DBO'] for row, value in st.session_state.ALM_DB[col].items()} for col in st.session_state.ALM_DB.keys()})
            df_NC = pd.DataFrame({col: {row: value['NC'] for row, value in st.session_state.ALM_DB[col].items()} for col in st.session_state.ALM_DB.keys()})
            df_EBP = pd.DataFrame({col: {row: value['EBP'] for row, value in st.session_state.ALM_DB[col].items()} for col in st.session_state.ALM_DB.keys()})
            # df_DBO= df_DBO[df_DBO.select_dtypes(include=['number']).columns].applymap('{:,.0f}'.format)
            # df_NC = df_NC[df_NC.select_dtypes(include=['number']).columns].applymap('{:,.0f}'.format)
            # df_EBP = df_EBP[df_EBP.select_dtypes(include=['number']).columns].applymap('{:,.0f}'.format)        
            
            selected_tab = st.selectbox("데이터를 선택하세요", ["DBO", "NC", "EBP"], index=0)
            if st.session_state.ALM_DB != {}:
            
            # 선택한 탭에 따라 데이터프레임을 표시
                rev_label = {       # rates_round(할인율) 레이블을 반전시키는 딕셔너리
                    'Min': 'Max',
                    'Q1': 'Q3',
                    'Median': 'Median',
                    'Q3': 'Q1',
                    'Max': 'Min',
                    '95% CI-': '95% CI+',
                    '95% CI+': '95% CI-',
                    'Mean': 'Mean'
                }
                if selected_tab == "DBO":
                    df_DBO_label = df_DBO.rename(columns=rev_label)
                    st.dataframe(df_DBO_label[df_DBO_label.select_dtypes(include=['number']).columns].applymap('{:,.0f}'.format))
                    # 그래프 레이블 및 타이틀 설정
                    fig, ax = plt.subplots()

                    for interval, interval_data in st.session_state.ALM_DB.items():
                        dbo_data = [entry['DBO'] for year, entry in interval_data.items()]
                        years = list(interval_data.keys())
                        label = rev_label.get(interval, f'{interval} DBO')
                        ax.plot(years, dbo_data, marker='o', linestyle='-', label=label)
                    
                    ax.set_xticks(range(min(years), max(years) + 1, 1))
                    formatter = ticker.FuncFormatter(millions_formatter)
                    ax.yaxis.set_major_formatter(formatter)
                    ax.set_xlabel('Year', fontsize=9)
                    ax.set_ylabel('DBO Value (Million)', fontsize=9)
                    ax.set_title('DBO Value by Year for Different Intervals')
                    ax.grid(True)
                    ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
                    st.session_state.DBO_plot = fig
                    st.pyplot(fig)
                            
                elif selected_tab == "NC":
                    df_NC_label = df_NC.rename(columns=rev_label)
                    st.dataframe(df_NC_label[df_NC_label.select_dtypes(include=['number']).columns].applymap('{:,.0f}'.format))
                    # 그래프 레이블 및 타이틀 설정
                    fig, ax = plt.subplots()
                    for interval, interval_data in st.session_state.ALM_DB.items():
                        nc_data = [entry['NC'] for year, entry in interval_data.items()]
                        years = list(interval_data.keys())
                        label = rev_label.get(interval, f'{interval} NC')
                        ax.plot(years, nc_data, marker='o', linestyle='-', label=label)
                    
                    ax.set_xticks(range(min(years), max(years) + 1, 1))
                    formatter = ticker.FuncFormatter(millions_formatter)
                    ax.yaxis.set_major_formatter(formatter)
                    ax.set_xlabel('Year', fontsize=9)
                    ax.set_ylabel('NC Value (Million)', fontsize=9)
                    ax.set_title('NC Value by Year for Different Intervals')
                    ax.grid(True)
                    ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
                    st.session_state.NC_plot = fig
                    st.pyplot(fig)
                            
                elif selected_tab == "EBP":
                    df_EBP_label = df_EBP.rename(columns=rev_label)                    
                    st.dataframe(df_EBP_label[df_EBP_label.select_dtypes(include=['number']).columns].applymap('{:,.0f}'.format))
                    # 그래프 레이블 및 타이틀 설정
                    fig, ax = plt.subplots()
                    for interval, interval_data in st.session_state.ALM_DB.items():
                        ebp_data = [entry['EBP'] for year, entry in interval_data.items()]
                        years = list(interval_data.keys())
                        label = rev_label.get(interval, f'{interval} EBP')
                        ax.plot(years, ebp_data, marker='o', linestyle='-', label=label)
                    
                    ax.set_xticks(range(min(years), max(years) + 1, 1))
                    formatter = ticker.FuncFormatter(millions_formatter)
                    ax.yaxis.set_major_formatter(formatter)
                    ax.set_xlabel('Year', fontsize=9)
                    ax.set_ylabel('EBP Value (Million)', fontsize=9)
                    ax.set_title('EBP Value by Year for Different Intervals')
                    ax.grid(True)
                    ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
                    st.session_state.EBP_plot = fig
                    st.pyplot(fig)        


        st.subheader("인별자료")
        # 다운로드 버튼을 누를 준비가 되었는지 확인하는 플래그
        generate_excel = False
        # '엑셀 파일 생성' 버튼 추가
        if st.button(label="엑셀 파일 생성", key = "DBO_ind_download"):
            generate_excel = True
        # 엑셀 파일 생성
        if generate_excel:            
            # ExcelWriter 객체 생성
            output = io.BytesIO()
            tabs = list(st.session_state.ALM_DB_ind.keys())
            
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                for tab in tabs:
                    for sim_yr in range(산출년수):
                        st.session_state.ALM_DB_ind[tab][sim_yr].to_excel(writer, sheet_name=f"{tab}_Year_{sim_yr}", index=True)
            excel_data = output.getvalue()
            # 다운로드 버튼 생성
            st.download_button(
                label=f"인별 데이터 내려받기",
                data=excel_data,
                file_name=f"ALM_DBO_indiv.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )    
        col5, col6, col7, col8 = st.columns(4)
        with col5:
            tabs = list(st.session_state.ALM_DB_ind.keys())
            selected_tab = st.selectbox("Choose an interval:", tabs)
        
            if selected_tab in tabs:
                data = st.session_state.ALM_DB_ind[selected_tab]
                # Create a selectbox for sim_yr within the selected interval
                sim_years = list(data.keys())
                with col6:
                    selected_sim_yr = st.selectbox("Choose a simulation year:", sim_years)
                with col7:
                    selected_mp = st.number_input("사번:", value = 1)
    
        if selected_tab in tabs:
            st.subheader(f"[{selected_tab}] Year_{selected_sim_yr}")
            # st.dataframe(data[selected_sim_yr].style.format({
            #     '근속년수':"{:.3f}",
            #     '지급률':"{:.3f}",
            #     '기준급여':"{:,.0f}",
            #     'DBO':"{:,.0f}",
            #     '추계액':"{:,.0f}",
            #     'NC':"{:,.0f}",
            #     'EBP':"{:,.0f}",
            #     't*DBO':"{:,.0f}",
            #     'DUR':"{:.3f}",
            #     '보정계수':"{:.3f}",
            # }))
            
            #Sample Test 데이터 display
            st.write(st.session_state.rates_round_filtered)

            st.subheader(f"사번: [{selected_mp}]")
            명부_dict = st.session_state.명부_dict
            cols = ['Min', 'Q1', 'Median', 'Q3', 'Max', '95% CI-', '95% CI+', 'Mean']
            column_names = ['시산연도', '시산연령', '경과기간', '임금상승률', '기준급여', '근속년수', '지급률', '생존확률', '퇴직률', '사망률', '퇴직급여_기간안분', '퇴직급여', 'v', 'DBO', 't*DBO', '퇴직급여_기간안분(기말)', 'NC', '추계액', 'EBP']
            int_rate = st.session_state.rates_round_filtered.iloc[selected_sim_yr, cols.index(selected_tab)]
            시산기준일 = (기준일 + pd.DateOffset(years = selected_sim_yr)).date()
            st.dataframe(pd.DataFrame(DBO(info_dict(명부_dict[selected_sim_yr], selected_mp - 1, 시산기준일, 0), 할인율=int_rate), columns=column_names).style.format({
                '시산연도':"{:.0f}",
                '시산연령':"{:.0f}",
                '경과기간':"{:.0f}",
                '임금상승률':"{:.3f}",
                '기준급여':"{:,.0f}",
                '근속년수':"{:.3f}",
                '지급률':"{:.3f}",
                '생존확률':"{:.4f}",
                '퇴직률':"{:.4f}",
                '사망률':"{:.4f}",
                '임금상승률':"{:.3f}",
                '퇴직급여_기간안분':"{:,.0f}",
                '퇴직급여':"{:,.0f}",
                'v':"{:.3f}",
                'DBO':"{:,.0f}",
                't*DBO':"{:,.0f}",
                '퇴직급여_기간안분(기말)':"{:,.0f}",   
                'NC':"{:,.0f}",
                '추계액':"{:,.0f}",
                'EBP':"{:,.0f}",       
            }))
            st.write(info_dict(명부_dict[selected_sim_yr], selected_mp - 1, 시산기준일, 0))
    with tab7:
        generate_excel = False
        
        sim_yr = range(산출년수)
        tabs = list(st.session_state.ALM_DB_ind.keys())
        
        ALM_DB_sense = {}
        ALM_DB_ind_sense = {}
        matrix_data = {}
        matrix_data_NC = {}
        if 'ALM_DB_sim' not in st.session_state:
            st.session_state.ALM_DB_sim = {}

        민감도_i = ['할인율-0.5%', '할인율-0.25%', '할인율', '할인율+0.25%', '할인율+0.5%']
        민감도_bu = ['bu-0.5%', 'bu-0.25%', 'bu', 'bu+0.25%', 'bu+0.5%']
        sense = [-0.005, -0.0025, 0, 0.0025, 0.005]
        for j in range(len(sense)):
            ALM_DB_sense[민감도_i[j]] = {}
            ALM_DB_ind_sense[민감도_i[j]] = {}    
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.write("### 민감도 및 지급시기 분포")
            selected_interval = st.selectbox("Choose an interval:", tabs, index = 3, key = 'DBO_sensitivity_interval')
            selected_sim_yr = st.selectbox("Choose a simulation year:", sim_yr, key = 'DBO_sensitivity_sim_yr')
        
        if st.button(label="민감도 산출", key = 'DBO_sensitivity'):
            할인율 = st.session_state.rates_round_filtered.iloc[selected_sim_yr][selected_interval]       
            cf = st.session_state.cf
            max_length = max(len(arr) for arr in cf[selected_interval][f'Year_{selected_sim_yr}'].values())
            sums = [0] * max_length
            for arr in cf[selected_interval][f'Year_{selected_sim_yr}'].values():
                for i in range(len(arr)):
                    sums[i] += arr[i]

            for j in range(len(sense)):    
                for k in range(len(sense)):
                    for sim_yr in range(0, 1):
                        시산기준일 = pd.to_datetime(date(기준일.year + sim_yr, 기준일.month, 기준일.day), format='%Y-%m-%d')
                        result_table, 명부, cf_data = simul(명부, 기준일, sim_yr, round(할인율 + sense[j], 4), round(sense[k], 4))
                        명부_dict[sim_yr] = 명부
                        total_emp = len(명부)
                        print(f"시산기준일: {시산기준일}, Discount Rate: {round(할인율 + sense[j], 4)}, bu: {민감도_bu[k]}, Year_{sim_yr}: Number of employees: {total_emp}")

                        # Save the result_table to the results_dict with key 'Year_0' or 'Year_1'
                        ALM_DB_sense[민감도_i[j]][민감도_bu[k]] = {
                            'DBO' : result_table['DBO'].sum(),
                            'NC' : result_table['NC'].sum(),
                            'EBP' : result_table['EBP'].sum(),
                        }
                        ALM_DB_ind_sense[민감도_i[j]][민감도_bu[k]] = result_table
                    명부 = 명부_dict[0]

            # --- 기존: matrix_data에는 DBO만 기록 ---
            for 할인율, bu_data in ALM_DB_sense.items():
                for bu, values in bu_data.items():
                    DBO = '{:,.0f}'.format(values['DBO'])
                    matrix_data.setdefault(할인율, {})[bu] = DBO

            # --- 추가: NC도 동일 방식으로 matrix_data_NC 생성 ---
            for 할인율, bu_data in ALM_DB_sense.items():
                for bu, values in bu_data.items():
                    NC = '{:,.0f}'.format(values['NC'])
                    matrix_data_NC.setdefault(할인율, {})[bu] = NC 


            # 기준값 추출 (예: 할인율과 bu 값이 모두 0인 경우의 DBO 값)
            baseline_DBO = ALM_DB_sense['할인율']['bu']['DBO']
            baseline_NC = ALM_DB_sense['할인율']['bu']['NC']

            # --- 오차율 계산 (DBO) ---
            error_rate_matrix = {
                할인율: {
                    bu: f"{((values['DBO'] - baseline_DBO) / baseline_DBO) * 100:.2f}%"
                    for bu, values in 할인율_data.items()
                }
                for 할인율, 할인율_data in ALM_DB_sense.items()
            }

            # --- 오차율 계산 (NC) ---
            error_rate_matrix_NC = {
                할인율: {
                    bu: f"{((values['NC'] - baseline_NC) / baseline_NC) * 100:.2f}%"
                    for bu, values in 할인율_data.items()
                }
                for 할인율, 할인율_data in ALM_DB_sense.items()
            }

            
            col1, col2, col3 = st.columns(3)

            with col1:
                st.write("### DBO 민감도")
                st.dataframe(pd.DataFrame(matrix_data))

            with col2:
                st.write("### NC 민감도")
                st.dataframe(pd.DataFrame(matrix_data_NC))

            with col3:
                st.write("### 연도별 DBO Flow")
                st.dataframe(pd.DataFrame(sums, columns=['연도별 DBO flow']), height=210)

            # ------- 다음 줄에 오차율 두 개 추가 -------
            st.write("### DBO 오차율(%)")
            st.dataframe(pd.DataFrame(error_rate_matrix))

            st.write("### NC 오차율(%)")
            st.dataframe(pd.DataFrame(error_rate_matrix_NC))

            # ---------------------
            #     엑셀 출력
            # ---------------------
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                pd.DataFrame(matrix_data).to_excel(writer, sheet_name='DBO 민감도', index=True)
                pd.DataFrame(matrix_data_NC).to_excel(writer, sheet_name='NC 민감도', index=True)

                pd.DataFrame(error_rate_matrix).to_excel(writer, sheet_name='DBO 민감도(%)', index=True)
                pd.DataFrame(error_rate_matrix_NC).to_excel(writer, sheet_name='NC 민감도(%)', index=True)

                pd.DataFrame(sums, columns=['연도별 DBO flow']).to_excel(
                    writer, sheet_name='연도별 DBO 분포', index=True
                )                   
            excel_data = output.getvalue()

            st.download_button(
                label="데이터 내려받기",
                data=excel_data,
                file_name="기타자료.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        
        st.write("### 명부 데이터")
        st.dataframe(명부)
        명부['임직원구분'] = 명부['가입대상분류'].map({1: '직원', 2: '임원', 3: '직군1', 4: '직군2'})

        colors = ["#31A2AC", "#FFC3A0","#ACECF7", "#AECDC2", "#FFAFBD", "#FFC3C2", "#F9E4E7", "#D4A5A5", "#392F5A",  "#61C0BF"]
        col1, col2 = st.columns(2)
        with col1: 
            fig1 = px.histogram(명부, x='연령', nbins=10, title='연령 분포',
                                color_discrete_sequence=colors,)
            fig1.update_yaxes(title_text='임직원 수')
            st.plotly_chart(fig1)
        with col2:
            # 근속년수 분포
            fig2 = px.histogram(명부, x='근속년수', nbins=10, title='근속년수 분포',
                                color_discrete_sequence=colors,)
            fig2.update_yaxes(title_text='임직원 수')
            st.plotly_chart(fig2)
        with col1:
            # 성별 분포
            fig3 = px.bar(명부['성별'].value_counts().reset_index(), x='성별', y='count', title='성별 분포',
                        color_discrete_sequence=colors,)
            fig3.update_yaxes(title_text='임직원 수')
            st.plotly_chart(fig3)
        with col2:
            # 임직원 분포
            fig4 = px.bar(명부['임직원구분'].value_counts().reset_index(), x='임직원구분', y='count', title='임직원 분포',
                        color_discrete_sequence=colors,)
            fig4.update_yaxes(title_text='임직원 수')
            st.plotly_chart(fig4)
        with col1:        
            # 평균임금 분포
            fig5 = px.scatter(명부, x='근속년수', y='기준급여', color='임직원구분', title='평균임금 분포', hover_data={'기준급여': ':,.0f'}
                            ,color_discrete_sequence=colors,)
            st.plotly_chart(fig5)
        
        
        if st.button(label="차트 이미지 생성", key = "chart_download"):
            generate_excel = True
            # Save Plotly figures as images
            fig1.write_image("fig1.png")
            fig2.write_image("fig2.png")
            fig3.write_image("fig3.png")
            fig4.write_image("fig4.png")
            fig5.write_image("fig5.png")

            # Create a new Excel workbook and add worksheets
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "연령 분포"
            ws1.sheet_view.showGridLines = False
            ws2 = wb.create_sheet(title="근속년수 분포")
            ws2.sheet_view.showGridLines = False
            ws3 = wb.create_sheet(title="성별 분포")
            ws3.sheet_view.showGridLines = False
            ws4 = wb.create_sheet(title="임직원 분포")
            ws4.sheet_view.showGridLines = False
            ws5 = wb.create_sheet(title="평균임금 분포")
            ws5.sheet_view.showGridLines = False

            # Insert images into Excel sheets
            img1 = Image("fig1.png")
            img2 = Image("fig2.png")
            img3 = Image("fig3.png")
            img4 = Image("fig4.png")
            img5 = Image("fig5.png")

            ws1.add_image(img1, 'A1')
            ws2.add_image(img2, 'A1')
            ws3.add_image(img3, 'A1')
            ws4.add_image(img4, 'A1')
            ws5.add_image(img5, 'A1')

            # Save the workbook
            wb.save("임직원 현황 차트.xlsx")        

        # 엑셀 파일 생성
        if generate_excel:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                correlation.to_excel(writer, sheet_name='corr', index=True)
                L.to_excel(writer, sheet_name='corr(cholesky)', index=True)
                명부['성별'].value_counts().reset_index().to_excel(writer, sheet_name='명부(성별)', index=False)
                명부['임직원구분'].value_counts().reset_index().to_excel(writer, sheet_name='명부(임직원구분)', index=False)
                명부['근속년수'].value_counts().sort_index().reset_index().to_excel(writer, sheet_name='명부(근속년수)', index=False)
                명부['연령'].value_counts().sort_index().reset_index().to_excel(writer, sheet_name='명부(연령)', index=False)
                pd.pivot_table(명부, values='기준급여', index='근속년수', columns='임직원구분', aggfunc='mean').to_excel(writer, sheet_name='명부(기준급여(평균))', index=True)
            excel_data = output.getvalue()

            # 다운로드 버튼 생성
            st.download_button(
                label="데이터 내려받기",
                data=excel_data,
                file_name="기타자료.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            generate_excel = False

    with tab8: 
        with st.form(key="자산군_설계_폼"):        
            col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
            with col1:  # 국내채권
                국내채권_설계 = asset_config_block(
                    title='국내채권',
                    base_default='KIS_total',
                    exp_key='Domestic_bond',
                    sigma_key='Domestic_bond',
                    model_default_index=1,
                    related_default=['Deposit_1Y','G_Bond_1Y','KIS_IR_2Y','KIS_IR_3Y','KIS_CY_3Y'],
                    key_prefix='국내채권',
                    use_index_suffix=True,   # 기초자산에 "_index" 붙임
                )

            with col2:  # 국내주식
                국내주식_설계 = asset_config_block(
                    title='국내주식',
                    base_default='KOSPI',
                    exp_key='Domestic_stock',
                    sigma_key='Domestic_stock',
                    model_default_index=1,
                    related_default=['KIS_total','MSCI_ACWI','Real_Estate','Deposit_1Y','G_Bond_10Y','Wage_rate'],
                    key_prefix='국내주식',
                    use_index_suffix=True,
                )

            with col3:  # 해외채권
                해외채권_설계 = asset_config_block(
                    title='해외채권',
                    base_default='ICE_GBMI',
                    exp_key='Global_bond',
                    sigma_key='Global_bond',
                    model_default_index=1,
                    related_default=['MSCI_ACWI','ICE_GCI','ICE_GGI','ICE_USCI_3Y','ICE_USCI_5Y'],
                    key_prefix='해외채권',
                    use_index_suffix=True,
                )

            with col4:  # 해외주식
                해외주식_설계 = asset_config_block(
                    title='해외주식',
                    base_default='MSCI_ACWI',
                    exp_key='Global_stock',
                    sigma_key='Global_stock',
                    model_default_index=1,
                    related_default=['ICE_GBMI','ICE_GCI','ICE_GGI','ICE_USCI_3Y','ICE_USCI_5Y'],
                    key_prefix='해외주식',
                    use_index_suffix=True,
                )

            with col5:  # 부동산
                부동산_설계 = asset_config_block(
                    title='부동산',
                    base_default='Real_Estate',
                    exp_key='Real_Estate',
                    sigma_key='Real_Estate',
                    model_default_index=0,
                    related_default=['CCI'],
                    key_prefix='부동산',
                    use_index_suffix=False,   # 부동산은 suffix 없이 원본 사용
                )

            with col6:  # PIGP 
                PIGP_설계 = asset_config_block(
                    title='PIGP',
                    base_default='Deposit_1Y',
                    exp_key='PIGP',
                    sigma_key='PIGP',
                    model_default_index=0,
                    related_default=['G_Bond_1Y','G_Bond_3Y','G_Bond_5Y','G_Bond_10Y'],
                    key_prefix='PIGP',
                    use_index_suffix=False,
                )

            with col7:  # 할인율(국채10년)
                Interest_rate_설계 = asset_config_block(
                    title='할인율(국채 10년)',
                    base_default='G_Bond_10Y',
                    exp_key='Interest_rate',
                    sigma_key='Interest_rate',
                    model_default_index=0,
                    related_default=['KIS_total','KOSPI','MSCI_ACWI','Real_Estate','Deposit_1Y','Wage_rate'],
                    key_prefix='Interest_rate',
                    use_index_suffix=False,
                )

            submitted = st.form_submit_button("저장")
            if submitted:
                st.session_state.자산설계 = pd.concat([
                    pd.DataFrame([국내채권_설계],   index=['Domestic_bond']),
                    pd.DataFrame([국내주식_설계],   index=['Domestic_stock']),
                    pd.DataFrame([해외주식_설계],   index=['Global_stock']),
                    pd.DataFrame([해외채권_설계],   index=['Global_bond']),
                    pd.DataFrame([부동산_설계],     index=['Real_Estate']),
                    pd.DataFrame([PIGP_설계],       index=['PIGP']),
                    pd.DataFrame([Interest_rate_설계], index=['Interest_rate']),
                ])
                st.toast("자산 설계를 저장했습니다.", icon="✅")
                # st.experimental_rerun()  # ← 저장 후 즉시 재실행
            
    with tab9:
        자산설계 = st.session_state.get('자산설계')
        if 자산설계 is None or (hasattr(자산설계, "empty") and 자산설계.empty):
            st.warning("tab8에서 '저장'을 눌러 자산설계를 먼저 저장해주세요.")
            st.stop()

        def arrow_compat(x):
            from numpy import ndarray
            if isinstance(x, (list, tuple, set, ndarray)):
                return ", ".join(map(str, list(x)))
            return x
        
        df_view = 자산설계.applymap(arrow_compat)
        st.dataframe(df_view)
        simulated_paths = {}
        quantile = {}
        ci = {}
        combined = {}
        real = {}
        kappa_dict = {}
        run_info = []
        
        # fl_return = st.file_uploader(":file_folder: Expected returns Upload",type=(["csv","xlsx","xls","pkl"]))
        # if fl_return is not None:
        #     Expected_returns = pickle.load(fl_return)
        #     st.dataframe(Expected_returns)        

        # fl_sigma = st.file_uploader(":file_folder: Expected sigma Upload",type=(["csv","xlsx","xls","pkl"]))
        # if fl_sigma is not None:
        #     Expected_sigma = pickle.load(fl_sigma)
        #     st.dataframe(Expected_sigma)

        # col1, col2, col3 = st.columns(3)
        # cols = [col1, col2, col3] 
        
        np.random.seed(seed)
        dt = 1 / 12
        index = range(simulation_times)
        columns = range(1, T + 1)
        
        asset_related = {
            asset: row['var_dw']
            for asset, row in 자산설계.iterrows()  
        }

        # 1) 전체 변수 목록 (중복 제거 + 순서 유지)
        var_dw = list(dict.fromkeys(v for vars_ in asset_related.values() for v in vars_))
        variable_number = len(var_dw)

        # 2) 상관행렬 (cor_data가 원시 시계열일 경우)
        correlation = cor_data[var_dw].corr()

        # 3) Cholesky 분해
        L = np.linalg.cholesky(correlation.values)

        # 4) 자산별 epsilon DataFrame 준비
        all_epsilon = {
            asset: pd.DataFrame(index=range(simulation_times), columns=range(1, T+1), dtype=float)
            for asset in asset_related
        }

        # 5) 시뮬레이션 타임 루프
        for x in range(1, T + 1):
            Z = np.random.normal(0, 1, size=(simulation_times, variable_number))
            correlated = Z @ L.T
            eps_df = pd.DataFrame(correlated, columns=var_dw)

            for asset, vars_ in asset_related.items():
                valid_vars = [c for c in vars_ if c in var_dw]
                if valid_vars:
                    all_epsilon[asset][x] = eps_df[valid_vars].mean(axis=1)
                else:
                    all_epsilon[asset][x] = np.nan


        #simulated_paths = st.session_state.simulated_paths
        if st.button(label="실행", key = '자산 시뮬레이션'):
            
            filename = "Asset_output.xlsx"  # 저장할 엑셀 파일의 이름
            writer = pd.ExcelWriter(filename, engine='xlsxwriter')
            
            for idx, asset in enumerate(list(asset_related.keys())):
                var = 자산설계.at[asset, '기초자산']
                sim_type = 자산설계.at[asset, '모델']
                mean = Expected_returns.at[f'Expected_returns',asset]
                sd = Expected_sigma.at[f'Expected_sigma',asset]

                r_expected  = float(Expected_returns.at[f'Expected_returns', asset])
                sd_expected = float(Expected_sigma.at[f'Expected_sigma',   asset])
                eps = all_epsilon[asset].values  # (simulation_times, T)

                # 공통 파라미터 사전 (자산별)
                params = {
                    "Asset": asset,
                    "Var": var,
                    "Model": sim_type,
                    "T": T,
                    "dt": dt,
                    "simulation_times": simulation_times,
                    "r_expected(annual)": r_expected,
                    "sd_expected": sd_expected,
                }

                if sim_type == 'Vasicek':
                    var = macro[var].dropna()
                    kappa0, theta0, sigma0, r0 = vasicek_calibrate(var, dt=dt)

                    theta = theta0
                    sigma = sd_expected  # 또는 sigma0 사용 가능
                    target_level = theta

                    res = minimize(
                        vasicek_objective, 
                        x0=max(kappa0, 1e-3),
                        args=(theta, sigma, r0, simulation_times, T, dt, eps, target_level),
                        method="SLSQP",
                        bounds = [(0.1, 0.5)]
                    )
                    kappa = float(res.x[0])
                    kappa_dict[asset] = kappa

                    paths = vasicek_simul(r0, kappa, theta, sigma, simulation_times, T, dt, eps)
                    simulated_paths[asset] = paths

                    # 모형별 상세 파라미터 채우기
                    params.update({
                        "r0(last)": r0,
                        "theta(level)": theta,
                        "sigma(level)": sigma,
                        "kappa": kappa,
                        "kappa0(calib)": kappa0,
                        "theta0(calib)": theta0,
                        "sigma0(calib)": sigma0,
                    })

                elif sim_type == 'GBM':
                    S0 = float(macro[var].iloc[-1])  # 지수 레벨 컬럼 사용 시: macro[f"{var}_index"].iloc[-1]
                    mu = find_optimal_mu(
                        target_annual_return=r_expected,
                        S0=S0, sigma=sd_expected,
                        simulation_times=simulation_times, T=T, dt=dt,
                        epsilon_var=eps,
                        init_mu=np.log(1 + r_expected) / 12
                    )
                    paths = gbm_simul(S0, mu, sd_expected, simulation_times, T, dt, eps)
                    simulated_paths[asset] = paths

                    params.update({
                        "S0(level)": S0,
                        "mu(monthly log)": mu,
                        "sigma(ret vol)": sd_expected,
                    })

                # elif sim_type == 'LSTM':
                #     #simulated_paths[asset] = LSTM_simulated_paths[asset]
                #     simulated_paths_list = []
                #     sim_train_predict_list = []
                #     sim_test_predict_list = []
                #     sim_train_x_values_list = []
                #     sim_test_x_values_list = []

                #     for _ in tqdm(range(10), desc=f"Simulating {asset}", unit="simulation"):
                #         # LSTM Simulation
                #         data, train_predict, test_predict, train_x_values, test_x_values, future_predictions, time_step = LSTM_simul(asset, seed)
                        
                #         # 각각의 결과를 리스트에 저장
                #         simulated_paths_list.append(future_predictions)
                #         sim_train_predict_list.append(train_predict)
                #         sim_test_predict_list.append(test_predict)
                #         sim_train_x_values_list.append(train_x_values)
                #         sim_test_x_values_list.append(test_x_values)
                        
                #         # Keras 모델의 메모리를 정리합니다.
                #         K.clear_session()
                #         # Python의 메모리를 정리합니다.
                #         gc.collect()

                #     # 최종적으로 각 리스트를 딕셔너리에 할당합니다.
                #     simulated_paths[asset] = simulated_paths_list     
                    
                var = st.session_state.자산설계.at[asset, '기초자산']
                render_and_cache_asset(asset, var, T, macro, simulated_paths, params) 
            st.success("시뮬레이션 결과를 세션에 저장했습니다.")

            plots = st.session_state.get("plots", {})
            params = st.session_state.get("params", {})        

            if not plots:
                st.info("아직 저장된 그래프가 없습니다. Tab9에서 '실행'을 먼저 눌러주세요.")
                st.stop()

        plots = st.session_state.get("plots", {})
        params = st.session_state.get("params", {})
        # 3개씩 나누어 출력
        cols = st.columns(3)
        for idx, (asset, fig) in enumerate(plots.items()):
            with cols[idx % 3]:   # 현재 행의 (0,1,2) 번째 column 중 하나
                st.pyplot(fig)
                st.caption(asset)
                if asset in params:
                    with st.expander(f"{asset} 파라미터 보기"):
                        st.json(params[asset])
            # 새 행으로 넘어가야 할 때 다시 st.columns(3) 생성
            if (idx + 1) % 3 == 0:
                cols = st.columns(3)
    
    with tab10:
        sim_paths = st.session_state.get('simulated_paths') or {}
        if not sim_paths or 'Interest_rate' not in sim_paths:
            st.warning("먼저 Tab9에서 '실행'을 눌러 시뮬레이션을 생성하세요.")
            st.stop()  # 여기서 즉시 중단        
        rates = pd.DataFrame(st.session_state.simulated_paths['Interest_rate']).T
        real = pd.concat([macro['G_Bond_10Y']] * simulation_times, axis=1)
        real.columns = range(real.shape[1])
        rates = pd.concat([real, rates], axis = 0, ignore_index=True)
        rates = rates.rolling(window=36).mean().dropna()
        rates = round((rates / 0.0025).round().clip(lower=0) * 0.0025, 4)
        date_range = pd.date_range(start=macro['Date'][0:].values[0], periods=len(macro['G_Bond_10Y']) + projection_period * 12, freq='M').strftime('%Y-%m-%d')
        rates = pd.concat([pd.DataFrame(date_range, columns = ['Date']), pd.DataFrame(rates)], axis=1)    
                
        st.write("### [자산배분용] 퇴직부채 시뮬레이션 산출")
        fl5 = st.file_uploader(":file_folder: DBO_sim Upload",type=(["csv","xlsx","xls","pkl"]))
        if fl5 is not None:
            ALM_DB_sim = pickle.load(fl5)
        else:
            if st.button(label="퇴직부채 산출(자산배분용)", key = 'DBO(for ALM)'):           
                strt = rates[rates['Date']==pd.to_datetime(기준일, format = '%Y-%m-%d').strftime('%Y-%m-%d')].index[0]
                for j in range(1, len(rates.columns)):
                    interval = rates.columns[j]
                    ALM_DB[interval] = {}
                    ALM_DB_ind[interval] = {}
                    results_dict[interval] = {}
                    cf[interval] = {}   
                    
                    for i in range(strt, strt + 산출년수 * 12, 12):
                        할인율.append(rates.iloc[i, j])
                            
                    for sim_yr in range(0, 산출년수):
                        시산기준일 = pd.to_datetime(date(기준일.year + sim_yr, 기준일.month, 기준일.day), format='%Y-%m-%d')
                        result_table, 명부, cf_data = simul(명부, 기준일, sim_yr, 할인율[(j - 1) * 산출년수 + sim_yr], 0)
                        명부_dict[sim_yr] = 명부
                        total_emp = len(명부)
                        print(f"시산기준일: {시산기준일}, Discount Rate: {할인율[(j - 1) * 산출년수 + sim_yr]}, interval: {interval}, Year_{sim_yr}: Number of employees: {total_emp}")

                        # Save the result_table to the results_dict with key 'Year_0' or 'Year_1'
                        results_dict[interval][f"Year_{sim_yr}"] = result_table
                        ALM_DB[interval][sim_yr] = {
                            'DBO' : results_dict[interval][f"Year_{sim_yr}"]['DBO'].sum(),
                            'NC' : results_dict[interval][f"Year_{sim_yr}"]['NC'].sum(),
                            'EBP' : results_dict[interval][f"Year_{sim_yr}"]['EBP'].sum(),
                        }
                    명부 = 명부_dict[0]
                
                pickle.dump(ALM_DB, open('ALM_DB_sim.pkl','wb'))        
                st.session_state.ALM_DB_sim = ALM_DB
            ALM_DB_sim = st.session_state.ALM_DB_sim
            
        if ALM_DB_sim != {}:
            #KEY 위치 바꾸기 ALM_DB[type][year][interval]
            new_ALM_DB = {}
            for key1, dict1 in ALM_DB_sim.items():
                for key2, value in dict1.items():
                    for main_key, val in value.items():
                        if main_key not in new_ALM_DB:
                            new_ALM_DB[main_key] = {}
                        if key2 not in new_ALM_DB[main_key]:
                            new_ALM_DB[main_key][key2] = {}
                        new_ALM_DB[main_key][key2][key1] = val

            col1, col2, col3 = st.columns(3)
            with col1:
                st.write("#### 할인율(국채 10년물)")
                rates_sorted = rates.sort_values(by='Date', ascending=False)
                st.dataframe(rates_sorted)
            with col2:
                st.write("#### 확정급여채무")
                col_labels = [f"Year{i} ({d})" for i, d in enumerate(date_labels)]
                tbl_DBO = pd.DataFrame(new_ALM_DB['DBO'])
                tbl_DBO.columns = col_labels  # 컬럼 이름 변경
                tbl_DBO.index.name = "Sim #"
                st.dataframe(tbl_DBO.style.format("{:,.0f}"))
            with col3:
                st.write("#### 확정급여채무 그래프")
                plt.figure(figsize=(10, 7))  # 그래프 크기 설정
                
                # 인터벌 별로 데이터 준비
                intervals = list(new_ALM_DB['DBO'][next(iter(new_ALM_DB['DBO']))].keys())
                years = list(new_ALM_DB['DBO'].keys())

                # 인터벌 별로 순회하며 그래프 추가
                for interval in intervals:
                    values = [new_ALM_DB['DBO'][year][interval] for year in years]
                    plt.plot(years, values, marker='o', linestyle='-', label=f'Interval {interval}')

                # 그래프 레이블 및 타이틀 설정
                plt.xlabel('Year')
                plt.ylabel('DBO Value')
                plt.title('DBO Simulation')
                plt.grid(True)
                plt.xticks(years)
                st.pyplot(plt)

        # 자산군 별 월별 시나리오를 연도별 시나리오로 구조 변환(1000 * 60 -> 1000 * 5)
        simulated_paths = st.session_state["simulated_paths"]
        combined_all = st.session_state["combined_all"]
        sim_paths_annual = {} 
        if simulated_paths != {}: 
            for idx, asset in enumerate(list(asset_related.keys())):
                # ① combined_df에서 macro+simulation 결합된 full 데이터 불러오기
                combined_all_filtered = combined_all[asset][combined_all[asset].index.month == 기준일.month]  # 기준일 동월 데이터만 추출
                # ③ 산출기준일 연도에 해당하는 값과 그 전년도 값 사용
                base_idx = combined_all_filtered.index.year.get_loc(기준일.year)

                if idx < 4:
                    # 0~3번 자산: 지수형 → 전년 대비 증감률
                    # 기준일 연도부터 이후 산출년도만큼 계산
                    returns_list = []
                    for i in range(base_idx, base_idx + 산출년수):
                        current_year = combined_all_filtered.iloc[i]
                        prev_year = combined_all_filtered.iloc[i - 1]
                        returns_list.append((current_year / prev_year) - 1)
                    
                # DataFrame으로 변환 (산출년수 행 × simulation_times 열)
                    sim_paths_annual[asset] = pd.DataFrame(returns_list, 
                                                        index=combined_all_filtered.index[base_idx:base_idx + 산출년수])
                else:
                    # 4번 이후 자산: 이미 연율값 → 그대로 사용
                    sim_paths_annual[asset] = combined_all_filtered.iloc[base_idx:base_idx + 산출년수]

            asset_list = list(sim_paths_annual.keys())
            selected_asset = st.selectbox("자산을 선택하세요", asset_list)

            # 선택한 자산의 연도별 시뮬레이션 결과 표시
            data = sim_paths_annual[selected_asset]
            date_str = data.index.strftime("%Y-%m")
            data.index = [f"Year {i} ({d})" for i, d in enumerate(date_str)]
            data.index.name = None
            st.write(f"▶ `{selected_asset}` 연도별 시뮬레이션 결과 (shape: {data.shape})") 
            st.dataframe(data)
        # ============================================
        # ALM 연도별 EF + Beam Search 최적화 (Streamlit 탭용)
        # -----------------------------
        # 0) 입력/환경
        # -----------------------------

        # 자산군 구분 및 순서
        LEVEL_ASSETS = ['Domestic_bond','Domestic_stock','Global_stock','Global_bond']  # 지수레벨
        YTM_ASSETS   = ['Real_Estate','PIGP']                                           # 레벨(YTM/지표)
        asset_order  = ['Domestic_bond','Domestic_stock','Global_bond','Global_stock','Real_Estate','PIGP']

         # -----------------------------
        # 3) 연도 반복 EF 곡선 (Streamlit 시각화)
        # -----------------------------

        # 1️⃣ 금리(rf) 확인 및 불러오기
        if ("rates_round" not in st.session_state) or \
        ("Mean" not in getattr(st.session_state["rates_round"], "columns", [])):
            st.error("세션에 rates_round['Mean']가 없습니다. 먼저 부채 산출(36M 평균 생성/반올림)을 실행하세요.")
            st.stop()

        rf_series = st.session_state.rates_round["Mean"]  # per-year rf 
        if len(rf_series) < 산출년수:
            st.warning(f"rf 연도 수({len(rf_series)})가 K_use({산출년수})보다 적어 부족 구간은 마지막 rf로 대체합니다.")
        
        # 2️⃣ MVO 제약 설정 UI
        st.subheader("MVO 제약 설정")
        col1, col2, col3, col4, col5, col6 = st.columns(6)
        with col1:
            domestic_bond_cap = st.slider("국내채권 상한(%)", min_value=0, max_value=60, value=60, step=1) / 100.0
        with col2:
            domestic_stock_cap = st.slider("국내주식 상한(%)", min_value=0, max_value=60, value=60, step=1) / 100.0
        with col3:
            global_bond_cap = st.slider("해외채권 상한(%)", min_value=0, max_value=60, value=15, step=1) / 100.0
        with col4:
            global_stock_cap = st.slider("해외주식 상한(%)", min_value=0, max_value=60, value=60, step=1) / 100.0
        with col5:
            real_estate_cap = st.slider("대체투자 상한(%)", min_value=0, max_value=60, value=15, step=1) / 100.0
        with col6:
            pigp_cap = st.slider("원리금보장 상한(%)", min_value=0, max_value=60, value=60, step=1) / 100.0

        # 자산별 상한 세션 저장
        st.session_state.asset_caps = {
            "Domestic_bond":  domestic_bond_cap,
            "Domestic_stock": domestic_stock_cap,
            "Global_bond":    global_bond_cap,
            "Global_stock":   global_stock_cap,
            "Real_Estate":    real_estate_cap,
            "PIGP":           pigp_cap,
        }

        A = len(asset_order)
        bounds = [(0.0, float(st.session_state.asset_caps.get(a, 1.0))) for a in asset_order]
        w0 = np.ones(A) / A

        # 3️⃣ 연도별 Efficient Frontier 계산
        ef_results = []
        R_years = [
            np.column_stack([sim_paths_annual[a].iloc[i, :].values for a in asset_order])
            for i in range(산출년수)
        ]
        for year in range(산출년수):
            wsY, tgY, muY, SigY, tarY = make_ef_for_year(R_years[year], bounds, n_target=20, w0=w0)
            ef_results.append({'wsY': wsY, 'muY': muY, 'SigY': SigY})

        # ============================================================
        # 4️⃣ Streamlit 시각화
        # ============================================================
        col1, col2 = st.columns(2)

        # --- (좌) EF 곡선 시각화 ---
        with col1:
            fig, ax = plt.subplots(figsize=(7, 5))
            colors = ['C0','C1','C2','C3','C4','C5','C6','C7']
            for i, d in enumerate(dates[1:], start=1):
                wsY, muY, SigY = ef_results[i]['wsY'], ef_results[i]['muY'], ef_results[i]['SigY']
                if wsY.size == 0:
                    continue
                #라벨
                fy_label = f"FY{d.strftime('%y')}"
                #EF계산
                risks = np.sqrt(np.einsum('ij,jk,ik->i', wsY, SigY, wsY))
                rets = wsY @ muY
                #EF곡선 그리기
                ax.plot(risks, rets, marker='o', ms=3, lw=1,
                        label=f"Year {i} ({fy_label})", 
                        color=colors[i % len(colors)])
            ax.set_xlabel("Volatility"); ax.set_ylabel("Expected Return")
            ax.xaxis.set_major_formatter(ticker.PercentFormatter(xmax=1.0, decimals=1))
            ax.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1.0, decimals=1))
            ax.set_title("Per-Year Efficient Frontiers")
            ax.grid(True)
            ax.legend()
            st.pyplot(fig)

        # --- (우) Sharpe_excess 기준 Top3 ---
        with col2:
            rows = [] 
            for i, d in enumerate(dates[1:], start=1):
                wsY, muY, SigY = ef_results[i]['wsY'], ef_results[i]['muY'], ef_results[i]['SigY']
                if wsY.size == 0:
                    continue

                target_date = pd.Timestamp(기준일) + pd.DateOffset(years=i)
                # rf_series에서 해당 날짜의 값 직접 추출
                rf = float(rf_series.loc[target_date])

                risks = np.sqrt(np.einsum('ij,jk,ik->i', wsY, SigY, wsY))
                rets = wsY @ muY
                sharpe_ex = np.divide(rets - rf, risks, out=np.zeros_like(rets), where=risks > 0)

                take = min(3, len(sharpe_ex))
                top_idx = np.argsort(sharpe_ex)[-take:][::-1]
                
                #라벨
                fy_label = f"FY{d.strftime('%y')}"
                
                for rank, k in enumerate(top_idx, start=1):
                    row = {
                        "Year": f"Year {i} ({fy_label})",
                        "Rank": rank,
                        "r_f": rf,
                        "Vol": float(risks[k]),
                        "ExpRet": float(rets[k]),
                        "Sharpe_excess": float(sharpe_ex[k]),
                    }
                    for a, w in zip(asset_order, wsY[k]):
                        row[a] = float(w)
                    rows.append(row)

            df_top = pd.DataFrame(rows)
            if not df_top.empty:
                df_top = df_top[["Year", "Rank", "r_f", *asset_order, "Vol", "ExpRet", "Sharpe_excess"]]

            st.write("**연도별 EF 최적 포트폴리오 Top 3 (Sharpe_excess 기준)**")
            st.dataframe(
                df_top.style.format(
                    {**{a: "{:.2%}" for a in asset_order},
                    "r_f": "{:.2%}", "Vol": "{:.2%}", "ExpRet": "{:.2%}", "Sharpe_excess": "{:.3f}"}
                ),
                use_container_width=True,
                height=560
            )

        # ------------------------------------------------------------
        # 4) ALM 자산배분 (연도별 가중치 최적화, 함수사용 최소화 버전)
        #  - 목표: min std(FR_T)
        #  - 제약: E[FR_T] ≥ 1, 전체평균 E[r] ≥ target,
        #         누적손실확률 ≤ p_neg_max, 각 해 sum(w_t)=1,
        #         0 ≤ w_{t,i} ≤ cap_i
        # ------------------------------------------------------------
        st.subheader("ALM 최적화 (연도별 비중)")

        colA, colB, colC = st.columns(3)
        with colA:
            initial_asset_input = st.number_input("적립금(직접입력)", step=1, value=0, key="초년도 적립금")
        with colB:
            target_return = st.number_input("목표 수익률(%)", value=3.0, step=0.5, format="%.1f")/100.0
        with colC:            
            p_neg_max     = st.number_input("전체기간 손실확률 상한(%)", value=10.0, step=1.0, format="%.1f")/100.0

        # ===========================
        # 4) ALM 자산배분 (연도별 최적화)
        # ===========================  

        if st.button("연도별 최적 비중 계산"):
            # 1️⃣ 기준연도 포함 연말 날짜 생성


            # 2️⃣ 자산 시뮬레이션 경로
            A3D = np.stack([sim_paths_annual[a].iloc[:산출년수, :].T.values for a in asset_order], axis=2) 

            # 3️⃣ 부채 행렬 (산출년수, S)
            new_ALM_DB = build_new_ALM_DB(ALM_DB_sim)
            EBP_mat = np.stack([np.asarray(new_ALM_DB['EBP'][k]).reshape(-1,) for k in range(산출년수)], axis=0)
            NC_mat  = np.stack([np.asarray(new_ALM_DB['NC'][k]).reshape(-1,)  for k in range(산출년수)], axis=0)
            DBO_mat = np.stack([np.asarray(new_ALM_DB['DBO'][k]).reshape(-1,) for k in range(산출년수)], axis=0)
            initial_asset_value = float(DBO_mat[0].mean()) if initial_asset_input == 0 else initial_asset_input        

            # 4️⃣ 자산 상한 및 초기 비중
            cap_vec = np.array([float(st.session_state.asset_caps.get(a, 1.0)) for a in asset_order], dtype=float)
            bounds = [(0.0, cap) for cap in np.tile(cap_vec, 산출년수)]
            x0 = np.full(산출년수 * A, 1.0 / A)

            # 6️⃣ 제약식 및 목적함수
            constraints = (
                [{'type': 'eq', 'fun': lambda x, t=t: np.sum(x.reshape(산출년수, A)[t, :]) - 1.0} for t in range(산출년수)] +
                [{'type': 'ineq', 'fun': lambda x: _compute_FR_trimmed(x, initial_asset_value)[2][:, -1].mean() - 1.0},
                {'type': 'ineq', 'fun': lambda x: _compute_FR_trimmed(x, initial_asset_value)[0].mean() - target_return},
                {'type': 'ineq', 'fun': lambda x: p_neg_max - ((np.prod(1 + _compute_FR_trimmed(x, initial_asset_value)[0], axis=1) - 1 < 0).mean())}]
            )

            obj = lambda x: np.std(_compute_FR_trimmed(x, initial_asset_value)[2][:, -1], ddof=1)   
                     
            opt = minimize(obj, x0, method="SLSQP", bounds=bounds,
                        constraints=constraints, options={"ftol": 1e-9, "maxiter": 2000})

            if not opt.success:
                st.error(f"[FAIL] {opt.message}")
            x_star = opt.x if opt.success else x0

            R_port_star, Asset_star, FR_star, Surp_star = _compute_FR(x_star, initial_asset_value)
            W_star = x_star.reshape(산출년수, A)
            st.session_state["R_port_star"] = R_port_star
            st.session_state["Asset_star"] = Asset_star
            st.session_state["FR_star"] = FR_star
            st.session_state["Surp_star"] = Surp_star
            st.session_state["W_star"] = W_star
            st.session_state["DBO_mat"] = DBO_mat
        
        if "R_port_star" in st.session_state:
            R_port_star = st.session_state["R_port_star"]
            Asset_star = st.session_state["Asset_star"]
            FR_star = st.session_state["FR_star"]
            Surp_star = st.session_state["Surp_star"]
            W_star = st.session_state["W_star"]
            DBO_mat = st.session_state["DBO_mat"]

            # ---- 결과 요약 ----
            W_star_adj = W_star[1:, :]
            dfW = pd.DataFrame(W_star_adj, index=[f"Year {i} ({d.strftime('%Y-%m')})" for i, d in enumerate(dates[1:], start=1)], columns=asset_order)
            st.subheader("연도별 최적 비중")
            st.dataframe(dfW.style.format("{:.2%}"))

            # ---- 성과 요약 ----
            R_port_adj = R_port_star[:, 1:]
            CumRet = np.cumprod(1 + R_port_adj, axis=1) - 1
            print("R_port_adj shape:", R_port_adj.shape)
            print("dates[1:] length:", len(dates[1:]))            
            dfCum = pd.DataFrame({
                "연도별 구간수익률(E[r_t])": R_port_adj.mean(axis=0),
                "연도별 원본손실 확률(Pr[r_t<0])": (R_port_adj < 0).mean(axis=0),
                "누적수익률(E[Cum(r_t)])": CumRet.mean(axis=0),
                "누적 표준편차(σ[Cum(r_t)])": CumRet.std(axis=0, ddof=1),
                "누적 원본손실 확률(Pr[Cum(r_t)<0])": (CumRet < 0).mean(axis=0),
            }, index=[f"Year {i} ({d.strftime('%Y-%m')})" for i, d in enumerate(dates[1:], start=1)])
            st.subheader("연도별 성과 요약")
            st.dataframe(dfCum.style.format("{:.2%}"))

            date_labels_adj = [d.strftime("%Y-%m") for d in dates[1:]]
            col1, col2 = st.columns(2)
            with col1:
                # 1️⃣ 누적수익률 플롯 (Cumulative Return Simulation, 95% 신뢰구간)
                fig, ax = plt.subplots(figsize=(8, 5))

                p10 = np.percentile(CumRet, 10, axis=0)
                p90 = np.percentile(CumRet, 90, axis=0)
                mean_ret = CumRet.mean(axis=0)

                ax.plot(date_labels_adj, CumRet.T, color='gray', alpha=0.25)
                ax.plot(date_labels_adj, CumRet.mean(axis=0), color='blue', alpha=0.7, lw=1.0, marker = 'o', markersize=2, label='Mean Path')
                ax.fill_between(date_labels_adj, p10, p90, color='blue', alpha=0.15, label='Percentile(10~90%)')

                # 데이터 레이블 추가
                for x, y in zip(date_labels_adj, mean_ret):
                    ax.text(
                        x, y, f"{y*100:.2f}%",          # 값 표시 (소수점 2째자리)
                        fontsize=8,
                        color='blue',
                        ha='center', va='bottom',   # 중앙 정렬, 점 위쪽에 표시
                        rotation=0,
                        fontweight='medium'
                    )
                ax.yaxis.set_major_formatter(ticker.PercentFormatter(1.0, decimals=1))
                ax.set_title("Cumulative Return Simulation")
                ax.set_ylabel("Cumulative Return")
                ax.grid(True, linestyle='--', alpha=0.5)
                ax.legend()
                st.pyplot(fig)

                # ----------------------------------------------------------

                # 2️⃣ 연도별 자산비중 영역그래프 (Stacked Area Chart)
                fig, ax = plt.subplots(figsize=(8, 5))
                W_cum = (W_star_adj.T ) * 100 # (A, 산출년수)

                ax.stackplot(date_labels_adj, W_cum, labels=asset_order, alpha=0.8)
                ax.set_title("Asset Allocation by Year")
                ax.set_ylabel("Weight(%)")
                ax.set_ylim(0, 100)
                ax.legend(loc='upper left', ncol=2, fontsize=8)
                st.pyplot(fig)

                # ----------------------------------------------------------
            with col2:
                # 3️⃣ DBO, Asset 평균 vs FR(라인) 그래프
                fig, ax1 = plt.subplots(figsize=(8, 5))

                # 평균값 계산
                DBO_mean = DBO_mat.mean(axis=1)[1:]
                Asset_mean = Asset_star.mean(axis=0)[1:]
                FR_mean = FR_star.mean(axis=0)[1:]

                bar_width = 0.35
                x = np.arange(len(date_labels_adj))

                ax1.bar(x - bar_width/2, DBO_mean, bar_width, label='DBO(Avg)', color='gray', alpha=0.6)
                ax1.bar(x + bar_width/2, Asset_mean, bar_width, label='Asset(Avg)', color='skyblue', alpha=0.8)
                ax1.set_ylabel("Amount(krw)")
                ax1.set_title("Asset/DBO Funded Ratio")
                ax1.set_xticks(x)
                ax1.set_xticklabels(date_labels_adj)
                ax1.grid(True, linestyle='--', alpha=0.4)

                ax2 = ax1.twinx()
                ax2.plot(x, FR_mean, color='brown', marker='o', markersize=2, label='FR(Avg)')
                ax2.set_ylabel("Funded Ratio")
                ax2.axhline(1.0, color='brown', linestyle='--', alpha=0.5)
                ax2.yaxis.set_major_formatter(ticker.PercentFormatter(1.0, decimals=1))

                # 범례 통합
                lines1, labels1 = ax1.get_legend_handles_labels()
                lines2, labels2 = ax2.get_legend_handles_labels()
                ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
                st.pyplot(fig)

                # ----------------------------------------------------------
                # 4️⃣ 연도별 Shortfall Risk (선택형)
                selected_label = st.selectbox("연도 선택", date_labels_adj, index=0)

                # 선택된 날짜에 대응하는 인덱스 찾기
                selected_idx = date_labels_adj.index(selected_label)

                # 선택한 연도의 누적수익률
                Cum_selected = CumRet[:, selected_idx]

                # 🎯 Shortfall Risk 계산 (누적수익률 < 0)
                shortfall_count = np.sum(Cum_selected < 0)
                shortfall_rate = shortfall_count / len(Cum_selected)

                fig, ax = plt.subplots(figsize=(8, 4))
                ax.hist(Cum_selected * 100, bins=50, color='skyblue', alpha=0.5, edgecolor='deepskyblue')
                ax.axvline(0, color='red', linestyle='--', lw=1)
                ax.set_title(
                    f"Year [{selected_idx+1}] {selected_label} Cumulative Return Dist.\n"
                    f"Shortfall Risk: {shortfall_rate*100:.3f}%  ({shortfall_count}/{len(Cum_selected)})"
                )
                ax.xaxis.set_major_formatter(ticker.FormatStrFormatter('%.1f'))
                ax.set_xlabel("Cumulative Return(%)")
                ax.set_ylabel("Frequency")
                st.pyplot(fig)

